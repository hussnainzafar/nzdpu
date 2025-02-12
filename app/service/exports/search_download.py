from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl.styles import Border, Font, Side
from pandas import ExcelWriter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import AggregatedObjectView, Choice, Config, Restatement
from app.db.redis import RedisClient
from app.routers.utils import get_choice_value, load_organization_by_lei
from app.schemas.enums import DefaultPromptsEnum, EmissionsUnitsEnum
from app.schemas.restatements import AttributePathsModel
from app.schemas.search import SearchQuery
from app.service.core.cache import CoreMemoryCache
from app.service.core.loaders import SubmissionLoader
from app.service.exports.forms_processor import (
    process_data_export,
    process_dataframe_data,
    process_financed_emissions,
    process_scope_3,
    process_scope_emissions,
    process_scope_ghg,
    process_target_validations,
    process_targets_and_progress,
)
from app.service.exports.headers.headers import (
    ExportOptions,
    SearchSheets,
    get_data_explorer_headers,
)
from app.service.exports.restatement import RestatementExportManager
from app.service.exports.utils import (
    extract_data_from_fe,
    format_datetime_for_downloads,
    get_column_names_financed_emissions,
    get_constraint_views,
    get_prompt_from_key,
    get_scope_emissions_data,
    load_choice_from_root_data,
)
from app.service.utils import (
    load_column_units,
    parse_and_transform_subscripts_to_normal,
    transform_subscript_to_normal,
)

logger = logging.getLogger("nzdpu_logger")

DATA_SOURCE_LIST = []


@dataclass
class SubmissionResult:
    nz_id: int
    submission: AggregatedObjectView | dict
    restatements: list[Restatement]


@dataclass
class SearchExportManager:
    """
    SearchExportManager class.

    Args:
        session (AsyncSession) - the SQLAlchemy session
        query_results (list) - the results from search-query
        query (SearchQuery) - the search query
        restatements_emissions (list) - list of restatements emissions
        restatements_targets (list) - list of restatements targets
        cache (RedisClient) - Redis client for caching
        default_single_headers (dict) - default headers for single values
        attributes_single_values (list) - list of attribute single values
        classification_information_toggler (bool) - toggler for classification information
        download_option (Enum) - toggler for download option
    """

    cache: RedisClient
    static_cache: CoreMemoryCache
    session: AsyncSession
    query_results: list[dict[str, Any]]
    query: SearchQuery = field(default_factory=SearchQuery)
    restatements_emissions: list = field(default_factory=list)
    restatements_targets: list = field(default_factory=list)
    default_single_headers: dict = field(default_factory=dict)
    attributes_single_values: list = field(default_factory=list)
    classification_information_toggler: bool = False
    download_option: str = ExportOptions.SEARCH.value

    async def _process_submission_rest(
        self, submission_loader: SubmissionLoader
    ) -> None:
        """
        Process submissions restatements and add three columns company_name,
        lei and reporting_year
        """
        ids = [sub.get("id") for sub in self.query_results]
        orgs = await self.static_cache.organizations()
        restatements_manager = RestatementExportManager(
            submissions=self.query_results,
            session=self.session,
            cache=self.cache,
            static_cache=self.static_cache,
            extract_targets=True,
        )

        restatements = (
            await restatements_manager.get_restatements_nz_id_mapping(ids)
        )
        restated_attr_paths_lookup = (
            restatements_manager.restated_attr_paths_mapping(restatements)
        )
        original_submissions = (
            await restatements_manager.get_original_submissions(
                restated_attr_paths_lookup, restatements
            )
        )
        for nz_id, rest_list in restatements.items():
            active_sub = restatements_manager.active_sub(rest_list)
            attr_paths = restated_attr_paths_lookup[nz_id]
            if attr_paths:
                constraint_views = await get_constraint_views(
                    self.session, attr_paths
                )
                for field_name, field_value in attr_paths.items():
                    await restatements_manager.extract_restated_field(
                        nz_id,
                        field_name,
                        field_value,
                        original_submissions,
                        active_sub,
                        rest_list,
                        constraint_views,
                        submission_loader,
                    )
        self.restatements_targets = restatements_manager.restatements_targets
        self.restatements_emissions = (
            restatements_manager.restatements_emissions
        )

    async def _process_attributes_fields_and_values(
        self, submission_loader: SubmissionLoader
    ) -> None:
        """
        Process attributes paths fields and get correct attribute name with
        description and value
        """
        default_prompts = {
            key.lower(): value.value
            for key, value in DefaultPromptsEnum.__members__.items()
        }
        # prevent from old data to stay
        self.default_single_headers = {}
        self.attributes_single_values = []
        self.classification_information_toggler = int(
            await self.classification_toggler()
        )
        for idx, result in enumerate(self.query_results, start=1):
            single_attr_values = {}
            for field in self.query.fields:
                # unpack attribute path
                attribute_path = AttributePathsModel.unpack_field_path(
                    str(field)
                )
                # unpack attribute path for _prompt field
                attribute_path_prompt = AttributePathsModel.unpack_field_path(
                    f"{field}_prompt"
                )
                # load attribute value
                try:
                    attribute_value = submission_loader.return_value(
                        attribute_path, result
                    )
                except KeyError:
                    attribute_value = "-"
                except (TypeError, Exception):
                    # when we hit type-error exception, that means
                    # form from this attribute is null, and we get
                    # value from form and place to that attribute
                    attribute_value = (
                        result[attribute_path.form]
                        if result[attribute_path.form]
                        else "-"
                    )
                # format true/false bool fields into yes/no
                if isinstance(attribute_value, bool):
                    attribute_value = (
                        "Yes"
                        if attribute_value
                        else "No"
                        if attribute_value is not None
                        else "-"
                    )
                # check attribute_value is multiple and format it
                if (
                    isinstance(attribute_value, list)
                    and len(attribute_value) > 1
                ):
                    new_value = ""
                    for counter, value in enumerate(attribute_value):
                        new_value += f"{value}"
                        if counter < len(attribute_value) - 1:
                            new_value += ", "
                    attribute_value = new_value
                # get attribute short description from _prompt field
                try:
                    attribute_desc = submission_loader.return_value(
                        attribute_path_prompt, result
                    )
                except (ValueError, Exception):
                    # for root fields we need to add exception and catch
                    # to format it differently
                    attribute_desc = await get_prompt_from_key(
                        session=self.session,
                        key=attribute_path.attribute,
                        values=result,
                    )
                    # check for default prompt which are not in DB
                    if not attribute_desc:
                        attribute_desc = default_prompts.get(
                            attribute_path.attribute, ""
                        )
                    # format if category is not assign to _prompt
                    if "{" in attribute_desc:
                        attribute_desc = re.sub(
                            r"\{.*?}",
                            str(attribute_path.choice.value),
                            attribute_desc,
                        )
                attribute_name = attribute_path.attribute
                # get attribute -> attribute_path.attribute
                # get attribute value -> attribute_value
                if idx == 1:
                    self.default_single_headers[attribute_name] = (
                        attribute_desc
                    )
                    # create units col for emissions attributes
                    if attribute_name.endswith(
                        tuple(EmissionsUnitsEnum.values())
                    ):
                        self.default_single_headers[
                            attribute_name + "_units"
                        ] = attribute_desc + " units"
                # add values for processing
                if (
                    attribute_name
                    in ["sics_sector", "sics_sub_sector", "sics_industry"]
                    and self.classification_information_toggler
                ):
                    single_attr_values[attribute_name] = (
                        "SICS classification information not available for"
                        " download. "
                    )
                else:
                    single_attr_values[attribute_name] = attribute_value
                # create units col for emissions attributes
                if attribute_name.endswith(tuple(EmissionsUnitsEnum.values())):
                    attribute_value_unit = await load_column_units(
                        attribute_path.attribute,
                        self.session,
                        self.static_cache,
                    )
                    single_attr_values[attribute_name + "_units"] = (
                        transform_subscript_to_normal(
                            attribute_value_unit[0]["actions"][0]["set"][
                                "units"
                            ]
                        )
                        if isinstance(attribute_value_unit, (dict, list))
                        else attribute_value_unit
                    )
            self.attributes_single_values.append(single_attr_values)

    async def _generate_excel_for_all(
        self,
        filename: str | None = None,
        source: bool = False,
        last_updated: bool = False,
    ) -> str:
        """
        Generate excel worksheets from data frames

        Returns:
            Excel filename for downloading
        """
        excel_filename = "Data_Explorer.xlsx" if filename is None else filename
        column_names_aum = await get_column_names_financed_emissions("aum")
        column_names_ge = await get_column_names_financed_emissions(
            "gross_exp"
        )
        thick_border_bottom = Border(bottom=Side(style="thin"))
        normal_border = Border()
        if source:
            # get source ids for mapping restated and source columns
            choice_list = await self.session.execute(
                select(Choice.choice_id).where(
                    Choice.set_name == "source_list"
                )
            )
            source_set = choice_list.scalars().all()
            DATA_SOURCE_LIST.extend(source_set)
        # create key
        key_data = [
            {
                "key": "-",
                "description": (
                    "Value has not been submitted by the source data"
                ),
            },
            {
                "key": "N/A",
                "description": (
                    "Value is expressed or deemed as non-applicable in "
                    "the source data"
                ),
            },
        ]
        metadata_headers = get_data_explorer_headers(
            SearchSheets.METADATA.value
        )
        scope_1_emissions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_1_EMISSIONS_SHEET.value
        )
        scope_1_ghg_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_1_GHG_BREAKDOWN_SHEET.value
        )
        scope_1_exclusions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_1_EXCLUSIONS_SHEET.value
        )
        scope_2_lb_emissions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_2_LB_EMISSIONS_SHEET.value
        )
        scope_2_lb_exclusions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_2_LB_EXCLUSIONS_SHEET.value
        )
        scope_2_mb_emissions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_2_MB_EMISSIONS_SHEET.value
        )
        scope_2_mb_exclusions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_2_MB_EXCLUSIONS_SHEET.value
        )
        scope_3_emissions_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_3_EMISSIONS_SHEET.value
        )
        scope_3_methodology_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_3_METHODOLOGY_SHEET.value
        )
        scope_3_exclusion_headers = get_data_explorer_headers(
            SearchSheets.SCOPE_3_EXCLUSIONS_SHEET.value
        )
        fe_aum_overview_headers = get_data_explorer_headers(
            SearchSheets.FE_OVERVIEW_AUM_SHEET.value
        )
        fe_aum_abs_headers = get_data_explorer_headers(
            SearchSheets.FE_ABSOLUTE_AUM_SHEET.value
        )
        fe_aum_intensity_headers = get_data_explorer_headers(
            SearchSheets.FE_INTENSITY_AUM_SHEET.value
        )
        fe_aum_dq_headers = get_data_explorer_headers(
            SearchSheets.FE_DATA_QUALITY_AUM_SHEET.value
        )
        fe_ge_overview_headers = get_data_explorer_headers(
            SearchSheets.FE_OVERVIEW_GE_SHEET.value
        )
        fe_ge_abs_headers = get_data_explorer_headers(
            SearchSheets.FE_ABSOLUTE_GE_SHEET.value
        )
        fe_ge_intensity_headers = get_data_explorer_headers(
            SearchSheets.FE_INTENSITY_GE_SHEET.value
        )
        fe_ge_dq_headers = get_data_explorer_headers(
            SearchSheets.FE_DATA_QUALITY_GE_SHEET.value
        )
        assure_verif_headers = get_data_explorer_headers(
            SearchSheets.AV_EMISSIONS_SHEET.value
        )
        restatements_e_headers = get_data_explorer_headers(
            SearchSheets.RESTATEMENTS_EMISSIONS_SHEET.value
        )
        targets_abs_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_ABSOLUTE_SHEET
        )
        targets_abs_progress_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_PROGRESS_ABSOLUTE_SHEET.value
        )
        targets_int_phys_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_INTENSITY_PHYS_SHEET.value
        )
        targets_int_econ_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_INTENSITY_ECON_SHEET.value
        )
        targets_int_phys_progress_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_PROGRESS_PHYS_INTENSITY_SHEET.value
        )
        targets_int_econ_progress_headers = get_data_explorer_headers(
            SearchSheets.TARGETS_PROGRESS_ECON_INTENSITY_SHEET.value
        )
        assure_verif_target = get_data_explorer_headers(
            SearchSheets.AV_TARGETS_SHEET.value
        )
        restatements_t_headers = get_data_explorer_headers(
            SearchSheets.RESTATEMENTS_TARGETS_SHEET.value
        )

        scope_1_emissions_head = [
            "total_scope_1_emissions_ghg",
            "total_scope_1_emissions_co2",
            "rationale_s1_non_disclose",
            "scope_1_methodology",
            "scope_1_change_type",
            "scope_1_change_description",
        ]
        fe_aum_overview_headers_comparison = [
            [
                "rationale_fn_aum_non_disclose",
                "fn_aum_total_emissions_ghg_sum",
                "fn_aum_total_emissions_co2_sum",
                "aum_coverage_total_perc",
                "fn_aum_exclusions",
                "currency_fn_aum",
                "fn_aum_change_type",
                "fn_aum_change_description",
            ]
        ]
        fe_gross_exp_overview_headers_comparison = [
            [
                "rationale_fn_aum_non_disclose",
                "fn_gross_exp_total_emissions_ghg_sum",
                "fn_gross_exp_total_emissions_co2_sum",
                "gross_exp_coverage_total_perc",
                "fn_gross_exp_exclusions",
                "currency_fn_gross_exp",
                "fn_gross_exp_change_type",
                "fn_gross_exp_change_description",
            ]
        ]
        # create key
        key_full = pd.DataFrame(key_data)
        key_full.columns = ["Key", "Description"]
        # create metadata
        metadata_full = pd.DataFrame(
            index=[0, 1], columns=list(metadata_headers.keys())
        )
        # restatements data frames
        restatements_e_df = pd.DataFrame(
            index=[0, 1], columns=list(restatements_e_headers.keys())
        )
        restatements_e_df.iloc[0] = [
            key.upper() for key in restatements_e_headers.keys()
        ]
        restatements_e_df.iloc[1] = list(restatements_e_headers.values())
        restatements_t_df = pd.DataFrame(
            index=[0, 1], columns=list(restatements_t_headers.keys())
        )
        restatements_t_df.iloc[0] = [
            key.upper() for key in restatements_t_headers.keys()
        ]
        restatements_t_df.iloc[1] = list(restatements_t_headers.values())
        emissions_res_df = pd.DataFrame(self.restatements_emissions)
        restatements_e_df = pd.concat(
            [restatements_e_df, emissions_res_df], ignore_index=True
        )
        targets_res_df = pd.DataFrame(self.restatements_targets)
        restatements_t_df = pd.concat(
            [restatements_t_df, targets_res_df], ignore_index=True
        )
        # create scopes data frames
        # create scope 1 emissions full
        scope_emissions_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_1_emissions_headers.keys()),
        )
        scope_emissions_full.iloc[0] = [
            key.upper() for key in scope_1_emissions_headers.keys()
        ]
        scope_emissions_full.iloc[1] = list(scope_1_emissions_headers.values())
        # create scope ghg breakdown full
        scope_ghg_full = pd.DataFrame(
            index=[0, 1], columns=list(scope_1_ghg_headers.keys())
        )
        scope_ghg_full.iloc[0] = [
            key.upper() for key in scope_1_ghg_headers.keys()
        ]
        scope_ghg_full.iloc[1] = list(scope_1_ghg_headers.values())
        # create scope 1 exclusions full
        scope_exclusion_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_1_exclusions_headers.keys()),
        )
        scope_exclusion_full.iloc[0] = [
            key.upper() for key in scope_1_exclusions_headers.keys()
        ]
        scope_exclusion_full.iloc[1] = list(
            scope_1_exclusions_headers.values()
        )
        # create scope 2 mb full
        # mb emissions
        scope_mb_emissions_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_2_mb_emissions_headers.keys()),
        )
        scope_mb_emissions_full.iloc[0] = [
            key.upper() for key in scope_2_mb_emissions_headers.keys()
        ]
        scope_mb_emissions_full.iloc[1] = list(
            scope_2_mb_emissions_headers.values()
        )
        # mb exclusions
        scope_mb_exclusion_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_2_mb_exclusions_headers.keys()),
        )
        scope_mb_exclusion_full.iloc[0] = [
            key.upper() for key in scope_2_mb_exclusions_headers.keys()
        ]
        scope_mb_exclusion_full.iloc[1] = list(
            scope_2_mb_exclusions_headers.values()
        )
        # create scope 2 lb full
        # lb emissions
        scope_lb_emissions_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_2_lb_emissions_headers.keys()),
        )
        scope_lb_emissions_full.iloc[0] = [
            key.upper() for key in scope_2_lb_emissions_headers.keys()
        ]
        scope_lb_emissions_full.iloc[1] = list(
            scope_2_lb_emissions_headers.values()
        )
        # lb exclusions
        scope_lb_exclusion_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_2_lb_exclusions_headers.keys()),
        )
        scope_lb_exclusion_full.iloc[0] = [
            key.upper() for key in scope_2_lb_exclusions_headers.keys()
        ]
        scope_lb_exclusion_full.iloc[1] = list(
            scope_2_lb_exclusions_headers.values()
        )
        # create scope 3 data frames
        # 3 emissions
        scope_3_emissions_full = pd.DataFrame(
            index=[0, 1],
            columns=list(scope_3_emissions_headers.keys()),
        )
        scope_3_emissions_full.iloc[0] = [
            key.upper() for key in scope_3_emissions_headers.keys()
        ]
        scope_3_emissions_full.iloc[1] = list(
            scope_3_emissions_headers.values()
        )
        scope_3_methodology_full = pd.DataFrame()
        scope_3_exclusion_full = pd.DataFrame()
        # 3 ghgp methodology
        methodology_ghgp_dataframes = {}
        for i in range(1, 17):
            if i == 16:
                dataframe_name = "scope_3_methodology_other_full"
            else:
                dataframe_name = f"scope_3_methodology_c{i}_full"
            methodology_ghgp_dataframes[dataframe_name] = pd.DataFrame()
        # 3 ghgp exclusion
        exclusion_ghgp_dataframes = {}
        for i in range(1, 17):
            if i == 16:
                dataframe_name = "scope_3_exclusion_other_full"
            else:
                dataframe_name = f"scope_3_exclusion_c{i}_full"
            exclusion_ghgp_dataframes[dataframe_name] = pd.DataFrame()
        # 3 iso methodology
        methodology_iso_dataframes = {}
        for i in range(1, 17):
            if i == 16:
                dataframe_name = "scope_3_methodology_other_full"
            else:
                dataframe_name = f"scope_3_methodology_c{i}_full"
            methodology_iso_dataframes[dataframe_name] = pd.DataFrame()
        # 3 iso exclusion
        exclusion_iso_dataframes = {}
        for i in range(1, 17):
            if i == 16:
                dataframe_name = "scope_3_exclusion_other_full"
            else:
                dataframe_name = f"scope_3_exclusion_c{i}_full"
            exclusion_iso_dataframes[dataframe_name] = pd.DataFrame()
        # create a&v data frames
        assure_verif_full = pd.DataFrame(
            index=[0, 1], columns=list(assure_verif_headers.keys())
        )
        assure_verif_full.iloc[0] = [
            key.upper() for key in assure_verif_headers.keys()
        ]
        assure_verif_full.iloc[1] = list(assure_verif_headers.values())
        assure_verif_targets_full = pd.DataFrame(
            index=[0, 1], columns=list(assure_verif_target.keys())
        )
        assure_verif_targets_full.iloc[0] = [
            key.upper() for key in assure_verif_target.keys()
        ]
        assure_verif_targets_full.iloc[1] = list(assure_verif_target.values())
        # create absolute targets df
        targets_abs_full = pd.DataFrame(
            index=[0, 1], columns=list(targets_abs_headers.keys())
        )
        targets_abs_full.iloc[0] = [
            key.upper() for key in targets_abs_headers.keys()
        ]
        targets_abs_full.iloc[1] = list(targets_abs_headers.values())
        targets_abs_progress_full = pd.DataFrame(
            index=[0, 1],
            columns=list(targets_abs_progress_headers.keys()),
        )
        targets_abs_progress_full.iloc[0] = [
            key.upper() for key in targets_abs_progress_headers.keys()
        ]
        targets_abs_progress_full.iloc[1] = list(
            targets_abs_progress_headers.values()
        )
        # create intensity targets df
        # phys intensity
        targets_int_phys_full = pd.DataFrame(
            index=[0, 1], columns=list(targets_int_phys_headers.keys())
        )
        targets_int_phys_full.iloc[0] = [
            key.upper() for key in targets_int_phys_headers.keys()
        ]
        targets_int_phys_full.iloc[1] = list(targets_int_phys_headers.values())
        # phys progress
        targets_int_phys_progress_full = pd.DataFrame(
            index=[0, 1],
            columns=list(targets_int_phys_progress_headers.keys()),
        )
        targets_int_phys_progress_full.iloc[0] = [
            key.upper() for key in targets_int_phys_progress_headers.keys()
        ]
        targets_int_phys_progress_full.iloc[1] = list(
            targets_int_phys_progress_headers.values()
        )
        # econ intensity
        targets_int_econ_full = pd.DataFrame(
            index=[0, 1],
            columns=list(targets_int_econ_headers.keys()),
        )
        targets_int_econ_full.iloc[0] = [
            key.upper() for key in targets_int_econ_headers.keys()
        ]
        targets_int_econ_full.iloc[1] = list(targets_int_econ_headers.values())
        # econ progress
        targets_int_econ_progress_full = pd.DataFrame(
            index=[0, 1],
            columns=list(targets_int_econ_progress_headers.keys()),
        )
        targets_int_econ_progress_full.iloc[0] = [
            key.upper() for key in targets_int_econ_progress_headers.keys()
        ]
        targets_int_econ_progress_full.iloc[1] = list(
            targets_int_econ_progress_headers.values()
        )
        # create financed emissions data frames
        # aum
        fe_aum_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_aum_overview_headers.keys()),
        )
        fe_aum_df.iloc[0] = [
            key.upper() for key in fe_aum_overview_headers.keys()
        ]
        fe_aum_df.iloc[1] = list(fe_aum_overview_headers.values())
        fn_aum_abs_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_aum_abs_headers.keys()),
        )
        fn_aum_abs_full_df.iloc[0] = [
            key.upper() for key in fe_aum_abs_headers.keys()
        ]
        fn_aum_abs_full_df.iloc[1] = list(fe_aum_abs_headers.values())
        fn_aum_int_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_aum_intensity_headers.keys()),
        )
        fn_aum_int_full_df.iloc[0] = [
            key.upper() for key in fe_aum_intensity_headers.keys()
        ]
        fn_aum_int_full_df.iloc[1] = list(fe_aum_intensity_headers.values())
        fn_aum_dq_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_aum_dq_headers.keys()),
        )
        fn_aum_dq_full_df.iloc[0] = [
            key.upper() for key in fe_aum_dq_headers.keys()
        ]
        fn_aum_dq_full_df.iloc[1] = list(fe_aum_dq_headers.values())
        # gross exp
        fe_ge_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_ge_overview_headers.keys()),
        )
        fe_ge_df.iloc[0] = [
            key.upper() for key in fe_ge_overview_headers.keys()
        ]
        fe_ge_df.iloc[1] = list(fe_ge_overview_headers.values())
        fn_ge_abs_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_ge_abs_headers.keys()),
        )
        fn_ge_abs_full_df.iloc[0] = [
            key.upper() for key in fe_ge_abs_headers.keys()
        ]
        fn_ge_abs_full_df.iloc[1] = list(fe_ge_abs_headers.values())
        fn_ge_int_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_ge_intensity_headers.keys()),
        )
        fn_ge_int_full_df.iloc[0] = [
            key.upper() for key in fe_ge_intensity_headers.keys()
        ]
        fn_ge_int_full_df.iloc[1] = list(fe_ge_intensity_headers.values())
        fn_ge_dq_full_df = pd.DataFrame(
            index=[0, 1],
            columns=list(fe_ge_dq_headers.keys()),
        )
        fn_ge_dq_full_df.iloc[0] = [
            key.upper() for key in fe_ge_dq_headers.keys()
        ]
        fn_ge_dq_full_df.iloc[1] = list(fe_ge_dq_headers.values())
        res: dict
        fe_df_mapping = {
            "aum_overview": {
                "df": fe_aum_df,
                "column_key": "fe_overview_headers",
                "res_key": None,
                "namespace": column_names_aum,
                "headers": True,
            },
            "ge_overview": {
                "df": fe_ge_df,
                "column_key": "fe_overview_headers",
                "res_key": None,
                "namespace": column_names_ge,
                "headers": True,
            },
            "aum_abs": {
                "df": fn_aum_abs_full_df,
                "column_key": "fn_coverage_headers",
                "res_key": "fn_aum_coverage",
                "namespace": column_names_aum,
                "headers": True,
            },
            "ge_abs": {
                "df": fn_ge_abs_full_df,
                "column_key": "fn_coverage_headers",
                "res_key": "fn_gross_exp_coverage",
                "namespace": column_names_ge,
                "headers": True,
            },
            "aum_int": {
                "df": fn_aum_int_full_df,
                "column_key": "fn_int_headers",
                "res_key": "fn_aum_intensity",
                "namespace": column_names_aum,
                "headers": True,
            },
            "ge_int": {
                "df": fn_ge_int_full_df,
                "column_key": "fn_int_headers",
                "res_key": "fn_gross_exp_intensity",
                "namespace": column_names_ge,
                "headers": True,
            },
            "aum_dq": {
                "df": fn_aum_dq_full_df,
                "column_key": "fn_dq_headers",
                "res_key": "fn_aum_data_quality",
                "namespace": column_names_aum,
                "headers": True,
            },
            "ge_dq": {
                "df": fn_ge_dq_full_df,
                "column_key": "fn_dq_headers",
                "res_key": "fn_gross_exp_data_quality",
                "namespace": column_names_ge,
                "headers": True,
            },
        }
        financial_lei = []
        # loop over results from search query
        for res in self.query_results:
            org = await load_organization_by_lei(
                res.get("lei"),
                self.session,
            )
            # add lei where organization is financial type
            if org.company_type == "Financial":
                financial_lei.append(org.lei)
            # get source value if we create excel data source
            source_value = res.get("source") if source else None
            last_updated_value = (
                format_datetime_for_downloads(res.get("reporting_datetime"))
                if last_updated
                else None
            )
            if source_value:
                source_value = await get_choice_value(
                    source_value, self.session, self.static_cache
                )
            unchanged_data = {
                "legal_entity_identifier": res.get("lei"),
                "company_name": res.get("company_name"),
                "data_model": await get_choice_value(
                    res.get("data_model"), self.session, self.static_cache
                ),
                "reporting_year": res.get("reporting_year"),
                "org_boundary_approach": (
                    last_updated_value
                    if source and last_updated
                    else (
                        await get_choice_value(
                            res.get("source"), self.session, self.static_cache
                        )
                        if source
                        and res.get("org_boundary_approach")
                        not in DATA_SOURCE_LIST
                        else await get_choice_value(
                            res.get("org_boundary_approach"),
                            self.session,
                            self.static_cache,
                        )
                    )
                ),
            }
            unchanged_desc_data = {
                "legal_entity_identifier": "Legal Entity Identifier (LEI)",
                "company_name": "Company name",
                "data_model": "Data model for disclosure",
                "reporting_year": "Reporting year",
                "org_boundary_approach": (
                    "Organizational boundary "
                    "approach used to consolidate "
                    "GHG emissions"
                ),
            }
            scope_root_data = await load_choice_from_root_data(
                unchanged_data, self.session, self.static_cache
            )
            date_start = format_datetime_for_downloads(
                res.get("date_start_reporting_year")
            )
            date_end = format_datetime_for_downloads(
                res.get("date_end_reporting_year")
            )
            # add metadata
            metadata = {
                "legal_entity_identifier": res.get("lei"),
                "company_name": res.get("company_name"),
                "company_jurisdiction": res.get("jurisdiction"),
                "sics_sector": (
                    "SICS classification information not available for"
                    " download. "
                ),
                "sics_sub_sector": (
                    "SICS classification information not available for"
                    " download. "
                ),
                "sics_industry": (
                    "SICS classification information not available for"
                    " download. "
                ),
                "data_model": await get_choice_value(
                    res.get("data_model"), self.session, self.static_cache
                ),
                "reporting_year": res.get("reporting_year"),
                "date_start_reporting_year": (
                    last_updated_value
                    if source and last_updated
                    else (
                        await get_choice_value(
                            res.get("source"), self.session, self.static_cache
                        )
                        if source
                        and res.get("org_boundary_approach")
                        not in DATA_SOURCE_LIST
                        else date_start
                    )
                ),
                "date_end_reporting_year": (
                    last_updated_value
                    if source and last_updated
                    else (
                        await get_choice_value(
                            res.get("source"), self.session, self.static_cache
                        )
                        if source
                        and res.get("org_boundary_approach")
                        not in DATA_SOURCE_LIST
                        else date_end
                    )
                ),
                "org_boundary_approach": (
                    last_updated_value
                    if source and last_updated
                    else (
                        await get_choice_value(
                            res.get("source"), self.session, self.static_cache
                        )
                        if source
                        and res.get("org_boundary_approach")
                        not in DATA_SOURCE_LIST
                        else await get_choice_value(
                            res.get("org_boundary_approach"),
                            self.session,
                            self.static_cache,
                        )
                    )
                ),
            }
            metadata_full.iloc[0] = [
                key.upper() for key in metadata_headers.keys()
            ]
            metadata_full.iloc[1] = list(metadata_headers.values())
            metadata_df = pd.DataFrame([metadata])
            metadata_full = pd.concat(
                [metadata_full, metadata_df], ignore_index=True
            )
            rationale = None
            rationale_other = None
            rationale_total = None
            rationale_desc = None
            rationale_other_desc = None
            rationale_total_desc = None
            key_in_list = any(
                key_scope in scope_1_emissions_head for key_scope in res
            )
            scope_emissions_data = {}
            scope_emissions_desc = {}
            restated_fields = []
            if key_in_list:
                # add scope 1 emissions to df
                scope_1_emissions_fields = get_scope_emissions_data(
                    res, "scope_1_emissions"
                )
                try:
                    (
                        scope_emissions_data,
                        scope_emissions_desc,
                    ) = await process_scope_emissions(
                        d=res,
                        emissions=scope_1_emissions_fields,
                        session=self.session,
                        source=source_value,
                        restated=restated_fields,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        download_option=self.download_option,
                        static_cache=self.static_cache,
                    )
                    # extract rationale for ghg breakdown
                    rationale = scope_emissions_data.pop(
                        "rationale_s1_ghg_bd_non_disclose", None
                    )
                    rationale_other = scope_emissions_data.pop(
                        "rationale_s1_ghg_bd_non_disclose_other", None
                    )
                    rationale_total = scope_emissions_data.pop(
                        "rationale_s1_ghg_not_equal_s1_total", None
                    )
                    rationale_desc = scope_emissions_desc.pop(
                        "rationale_s1_ghg_bd_non_disclose", None
                    )
                    rationale_other_desc = scope_emissions_desc.pop(
                        "rationale_s1_ghg_bd_non_disclose_other", None
                    )
                    rationale_total_desc = scope_emissions_desc.pop(
                        "rationale_s1_ghg_not_equal_s1_total", None
                    )
                except KeyError:
                    pass
            scope_emissions_full = await process_dataframe_data(
                data=scope_emissions_data,
                desc=scope_emissions_desc,
                df_full=scope_emissions_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 1 ghg breakdown to df
            scope_ghg_data_new = {}
            scope_ghg_desc_new = {}
            scope_1_ghgb = res.get("scope_1_ghg_breakdown", [])
            if scope_1_ghgb:
                try:
                    (
                        scope_ghg_data,
                        scope_desc,
                    ) = await process_scope_ghg(
                        scope_ghg_list=scope_1_ghgb,
                        session=self.session,
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        static_cache=self.static_cache,
                    )
                    scope_ghg_data_new = {
                        **{"rationale_s1_ghg_bd_non_disclose": rationale},
                        **{
                            "rationale_s1_ghg_bd_non_disclose_other": (
                                rationale_other
                            )
                        },
                        **{
                            "rationale_s1_ghg_not_equal_s1_total": (
                                rationale_total
                            )
                        },
                        **scope_ghg_data,
                    }
                    scope_ghg_desc_new = {
                        **{
                            "rationale_s1_ghg_bd_non_disclose": (
                                rationale_desc
                            )
                        },
                        **{
                            "rationale_s1_ghg_bd_non_disclose_other": (
                                rationale_other_desc
                            )
                        },
                        **{
                            "rationale_s1_ghg_not_equal_s1_total": (
                                rationale_total_desc
                            )
                        },
                        **scope_desc,
                    }
                except KeyError:
                    pass
            scope_ghg_full = await process_dataframe_data(
                data=scope_ghg_data_new,
                desc=scope_ghg_desc_new,
                df_full=scope_ghg_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 1 exclusion to df
            scope_exclusion_data = {}
            scope_exclusion_desc = {}
            scope_1_exc = res.get("scope_1_exclusion", [])
            if scope_1_exc:
                try:
                    (
                        scope_exclusion_data,
                        scope_exclusion_desc,
                    ) = await process_data_export(
                        data_export_list=scope_1_exc,
                        form_name="scope_1_exclusion",
                        session=self.session,
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        download_option=self.download_option,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            scope_exclusion_full = await process_dataframe_data(
                data=scope_exclusion_data,
                desc=scope_exclusion_desc,
                df_full=scope_exclusion_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 2 mb emissions to df
            scope_mb_emissions_data = {}
            scope_mb_emissions_desc = {}
            scope_2_mb_emissions_fields = get_scope_emissions_data(
                res, "scope_2_mb_emissions"
            )
            try:
                (
                    scope_mb_emissions_data,
                    scope_mb_emissions_desc,
                ) = await process_scope_emissions(
                    d=res,
                    emissions=scope_2_mb_emissions_fields,
                    session=self.session,
                    source=source_value,
                    last_updated=last_updated_value,
                    restated=restated_fields,
                    data_source_list=DATA_SOURCE_LIST,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                )
            except KeyError:
                pass
            scope_mb_emissions_full = await process_dataframe_data(
                data=scope_mb_emissions_data,
                desc=scope_mb_emissions_desc,
                df_full=scope_mb_emissions_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 2 mb exclusion to df
            scope_mb_exclusion_data = {}
            scope_mb_exclusion_desc = {}
            scope_2_mbx = res.get("scope_2_mb_exclusion", [])
            if scope_2_mbx:
                try:
                    (
                        scope_mb_exclusion_data,
                        scope_mb_exclusion_desc,
                    ) = await process_data_export(
                        data_export_list=scope_2_mbx,
                        form_name="scope_2_mb_exclusion",
                        session=self.session,
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        download_option=self.download_option,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            scope_mb_exclusion_full = await process_dataframe_data(
                data=scope_mb_exclusion_data,
                desc=scope_mb_exclusion_desc,
                df_full=scope_mb_exclusion_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 2 lb emissions to df
            scope_lb_emissions_data = {}
            scope_lb_emissions_desc = {}
            scope_2_lb_emissions_fields = get_scope_emissions_data(
                res, "scope_2_lb_emissions"
            )
            try:
                (
                    scope_lb_emissions_data,
                    scope_lb_emissions_desc,
                ) = await process_scope_emissions(
                    d=res,
                    emissions=scope_2_lb_emissions_fields,
                    session=self.session,
                    source=source_value,
                    restated=restated_fields,
                    last_updated=last_updated_value,
                    data_source_list=DATA_SOURCE_LIST,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                )
            except KeyError:
                pass
            scope_lb_emissions_full = await process_dataframe_data(
                data=scope_lb_emissions_data,
                desc=scope_lb_emissions_desc,
                df_full=scope_lb_emissions_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 2 lb exclusion to df
            scope_lb_exclusion_data = {}
            scope_lb_exclusion_desc = {}
            scope_2_lbx = res.get("scope_2_lb_exclusion", [])
            if scope_2_lbx:
                try:
                    (
                        scope_lb_exclusion_data,
                        scope_lb_exclusion_desc,
                    ) = await process_data_export(
                        data_export_list=scope_2_lbx,
                        form_name="scope_2_lb_exclusion",
                        session=self.session,
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        download_option=self.download_option,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            scope_lb_exclusion_full = await process_dataframe_data(
                data=scope_lb_exclusion_data,
                desc=scope_lb_exclusion_desc,
                df_full=scope_lb_exclusion_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add scope 3 ghg emissions to df
            scope_source = {
                "scope_3_breakdown_source": res.get(
                    "scope_3_breakdown_source", None
                ),
                "total_scope_3_emissions_ghg": res.get(
                    "total_scope_3_emissions_ghg", None
                ),
                "total_scope_3_emissions_co2": res.get(
                    "total_scope_3_emissions_co2", None
                ),
            }
            scope_3_ghgp_emissions_data = {}
            scope_3_ghgp_methodology_data = {}
            scope_3_ghgp_exclusion_data = {}
            scope_3_ghgp_emissions_desc = {}
            scope_3_ghgp_methodology_desc = {}
            scope_3_ghgp_exclusion_desc = {}
            scope_3_ghgp = res.get("scope_3_ghg_protocol", [])
            if scope_3_ghgp:
                try:
                    (
                        scope_3_ghgp_emissions_data,
                        scope_3_ghgp_methodology_data,
                        scope_3_ghgp_exclusion_data,
                        scope_3_ghgp_emissions_desc,
                        scope_3_ghgp_methodology_desc,
                        scope_3_ghgp_exclusion_desc,
                    ) = await process_scope_3(
                        scope_list=scope_3_ghgp,
                        scope_source=scope_source,
                        scope_group="ghgp",
                        session=self.session,
                        form_name="scope_3_ghg_protocol",
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            scope_3_iso_emissions_data = {}
            scope_3_iso_methodology_data = {}
            scope_3_iso_exclusion_data = {}
            scope_3_iso_emissions_desc = {}
            scope_3_iso_methodology_desc = {}
            scope_3_iso_exclusion_desc = {}
            scope_3_iso = res.get("scope_3_iso", [])
            if scope_3_iso:
                # add scope 3 iso emission to df
                try:
                    (
                        scope_3_iso_emissions_data,
                        scope_3_iso_methodology_data,
                        scope_3_iso_exclusion_data,
                        scope_3_iso_emissions_desc,
                        scope_3_iso_methodology_desc,
                        scope_3_iso_exclusion_desc,
                    ) = await process_scope_3(
                        scope_list=scope_3_iso,
                        scope_source=scope_source,
                        scope_group="iso",
                        session=self.session,
                        form_name="scope_3_iso",
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            scope_3_emissions_data = {
                **scope_3_ghgp_emissions_data,
                **scope_3_iso_emissions_data,
            }
            scope_3_emissions_desc = {
                **scope_3_ghgp_emissions_desc,
                **scope_3_iso_emissions_desc,
            }
            # add scope 3 emissions to df
            if scope_3_emissions_data:
                scope_3_emissions_full = await process_dataframe_data(
                    data=scope_3_emissions_data,
                    desc=scope_3_emissions_desc,
                    df_full=scope_3_emissions_full,
                    unchanged_data=unchanged_data,
                    unchanged_desc_data=unchanged_desc_data,
                )
            if scope_3_ghgp_methodology_data or scope_3_iso_methodology_data:
                scope_3_methodology_full = await process_dataframe_data(
                    data={},
                    desc={},
                    df_full=scope_3_methodology_full,
                    unchanged_data=unchanged_data,
                    unchanged_desc_data=unchanged_desc_data,
                )
            # add scope 3 ghgp methodology to df
            if scope_3_ghgp_methodology_data:
                for i in range(1, 17):
                    key_data = f"scope_3_c{i}_data"
                    key_desc = f"scope_3_c{i}_desc"

                    if i == 16:
                        key_data = "scope_3_other_data"
                        key_desc = "scope_3_other_desc"

                    key_scope_full = (
                        f"scope_3_methodology_c{i}_full"
                        if i != 16
                        else "scope_3_methodology_other_full"
                    )
                    if scope_3_ghgp_methodology_data.get(key_data):
                        methodology_ghgp_dataframes[
                            key_scope_full
                        ] = await process_dataframe_data(
                            data=scope_3_ghgp_methodology_data.get(key_data),
                            desc=scope_3_ghgp_methodology_desc.get(key_desc),
                            df_full=methodology_ghgp_dataframes.get(
                                key_scope_full, pd.DataFrame()
                            ),
                        )
            if scope_3_iso_methodology_data:
                for i in range(1, 17):
                    key_data = f"scope_3_c{i}_data"
                    key_desc = f"scope_3_c{i}_desc"

                    if i == 16:
                        key_data = "scope_3_other_data"
                        key_desc = "scope_3_other_desc"

                    key_scope_full = (
                        f"scope_3_methodology_c{i}_full"
                        if i != 16
                        else "scope_3_methodology_other_full"
                    )
                    if scope_3_iso_methodology_data.get(key_data):
                        methodology_ghgp_dataframes[
                            key_scope_full
                        ] = await process_dataframe_data(
                            data=scope_3_iso_methodology_data.get(key_data),
                            desc=scope_3_iso_methodology_desc.get(key_desc),
                            df_full=methodology_iso_dataframes.get(
                                key_scope_full, pd.DataFrame()
                            ),
                        )

            if scope_3_ghgp_exclusion_data or scope_3_iso_exclusion_data:
                scope_3_exclusion_full = await process_dataframe_data(
                    data={},
                    desc={},
                    df_full=scope_3_exclusion_full,
                    unchanged_data=unchanged_data,
                    unchanged_desc_data=unchanged_desc_data,
                )
            # add scope 3 ghgp exclusion to df
            if scope_3_ghgp_exclusion_data:
                for i in range(1, 17):
                    key_data = f"scope_3_c{i}_data"
                    key_desc = f"scope_3_c{i}_desc"

                    if i == 16:
                        key_data = "scope_3_other_data"
                        key_desc = "scope_3_other_desc"

                    key_scope_full = (
                        f"scope_3_exclusion_c{i}_full"
                        if i != 16
                        else "scope_3_exclusion_other_full"
                    )
                    if scope_3_ghgp_exclusion_data.get(key_data):
                        exclusion_ghgp_dataframes[
                            key_scope_full
                        ] = await process_dataframe_data(
                            data=scope_3_ghgp_exclusion_data.get(key_data),
                            desc=scope_3_ghgp_exclusion_desc.get(key_desc),
                            df_full=exclusion_ghgp_dataframes.get(
                                key_scope_full, pd.DataFrame()
                            ),
                        )
                # add scope 3 iso methodology to df
                for i in range(1, 17):
                    key_data = f"scope_3_c{i}_data"
                    key_desc = f"scope_3_c{i}_desc"

                    if i == 16:
                        key_data = "scope_3_other_data"
                        key_desc = "scope_3_other_desc"

                    key_scope_full = (
                        f"scope_3_methodology_c{i}_full"
                        if i != 16
                        else "scope_3_methodology_other_full"
                    )
                    if scope_3_iso_methodology_data.get(key_data):
                        methodology_iso_dataframes[
                            key_scope_full
                        ] = await process_dataframe_data(
                            data=scope_3_iso_methodology_data.get(key_data),
                            desc=scope_3_iso_methodology_desc.get(key_desc),
                            df_full=methodology_iso_dataframes.get(
                                key_scope_full, pd.DataFrame()
                            ),
                        )
            # add scope 3 iso exclusion to df
            if scope_3_iso_exclusion_data:
                for i in range(1, 17):
                    key_data = f"scope_3_c{i}_data"
                    key_desc = f"scope_3_c{i}_desc"

                    if i == 16:
                        key_data = "scope_3_other_data"
                        key_desc = "scope_3_other_desc"

                    key_scope_full = (
                        f"scope_3_exclusion_c{i}_full"
                        if i != 16
                        else "scope_3_exclusion_other_full"
                    )
                    if scope_3_iso_exclusion_data.get(key_data):
                        exclusion_iso_dataframes[
                            key_scope_full
                        ] = await process_dataframe_data(
                            data=scope_3_iso_exclusion_data.get(key_data),
                            desc=scope_3_iso_exclusion_desc.get(key_desc),
                            df_full=exclusion_iso_dataframes.get(
                                key_scope_full, pd.DataFrame()
                            ),
                        )
            # add v&a emissions to df
            assure_verif_data = {}
            assure_verif_desc = {}
            assure = res.get("assure_verif_valid_statement_dis", [])
            if assure:
                try:
                    rationale_assure = res.get(
                        "rationale_emissions_assure_verif_non_disclose", None
                    )
                    (
                        assure_verif_data,
                        assure_verif_desc,
                    ) = await process_data_export(
                        data_export_list=assure,
                        form_name="assure_verif_valid_statement_dis",
                        session=self.session,
                        source=source_value,
                        last_updated=last_updated_value,
                        data_source_list=DATA_SOURCE_LIST,
                        download_option=self.download_option,
                        rationale_assure=rationale_assure,
                        static_cache=self.static_cache,
                    )
                except KeyError:
                    pass
            assure_verif_full = await process_dataframe_data(
                data=assure_verif_data,
                desc=assure_verif_desc,
                df_full=assure_verif_full,
                unchanged_data=unchanged_data,
                unchanged_desc_data=unchanged_desc_data,
            )
            # add v&a targets to df
            assure_verif_target_data = {}
            assure_verif_target_desc = {}
            assure_tgt = res.get("assure_verif_statement_target", [])
            if assure_tgt:
                for target_valid in assure_tgt:
                    if isinstance(target_valid, dict):
                        try:
                            (
                                assure_verif_target_data,
                                assure_verif_target_desc,
                            ) = await process_target_validations(
                                target_validation=target_valid,
                                rationale_target=res.get(
                                    "rationale_target_valid_non_disclose",
                                    None,
                                ),
                                session=self.session,
                                source=source_value,
                                last_updated=last_updated_value,
                                data_source_list=DATA_SOURCE_LIST,
                                static_cache=self.static_cache,
                            )
                        except KeyError:
                            pass
                    assure_verif_targets_full = await process_dataframe_data(
                        data=assure_verif_target_data,
                        desc=assure_verif_target_desc,
                        df_full=assure_verif_targets_full,
                        unchanged_data=unchanged_data,
                        unchanged_desc_data=unchanged_desc_data,
                    )
            # add absolute targets to df
            tgt_abs = res.get("target_abs", [])
            if tgt_abs:
                for target_abs in tgt_abs:
                    if isinstance(target_abs, dict):
                        try:
                            (
                                targets_abs_data,
                                targets_abs_progress_data,
                                targets_abs_desc,
                                targets_abs_progress_desc,
                            ) = await process_targets_and_progress(
                                targets=target_abs,
                                target="abs_",
                                session=self.session,
                                source=source_value,
                                last_updated=last_updated,
                                data_source_list=DATA_SOURCE_LIST,
                                static_cache=self.static_cache,
                            )
                        except KeyError:
                            targets_abs_data = {}
                            targets_abs_progress_data = {}
                            targets_abs_desc = {}
                            targets_abs_progress_desc = {}
                    else:
                        targets_abs_data = {}
                        targets_abs_progress_data = {}
                        targets_abs_desc = {}
                        targets_abs_progress_desc = {}
                    targets_abs_full = await process_dataframe_data(
                        data=targets_abs_data,
                        desc=targets_abs_desc,
                        df_full=targets_abs_full,
                        unchanged_data=unchanged_data,
                        unchanged_desc_data=unchanged_desc_data,
                    )
                    # add absolute targets progress to df
                    targets_abs_progress_full = await process_dataframe_data(
                        data=targets_abs_progress_data,
                        desc=targets_abs_progress_desc,
                        df_full=targets_abs_progress_full,
                        unchanged_data=unchanged_data,
                        unchanged_desc_data=unchanged_desc_data,
                    )
            # add intensity targets (physical and economic) to df
            tgt_int = res.get("target_int", [])
            if tgt_int:
                for target_int in tgt_int:
                    if isinstance(target_int, dict):
                        try:
                            (
                                targets_int_data,
                                targets_int_progress_data,
                                targets_int_desc,
                                targets_int_progress_desc,
                            ) = await process_targets_and_progress(
                                targets=target_int,
                                target="int_",
                                session=self.session,
                                source=source_value,
                                last_updated=last_updated,
                                data_source_list=DATA_SOURCE_LIST,
                                static_cache=self.static_cache,
                            )
                        except KeyError:
                            targets_int_data = {}
                            targets_int_progress_data = {}
                            targets_int_desc = {}
                            targets_int_progress_desc = {}
                    else:
                        targets_int_data = {}
                        targets_int_progress_data = {}
                        targets_int_desc = {}
                        targets_int_progress_desc = {}
                    # add targets intensity physical
                    targets_int_phys_full = await process_dataframe_data(
                        data=targets_int_data.get("target_int_physical", {}),
                        desc=targets_int_desc.get("target_int_physical", {}),
                        df_full=targets_int_phys_full,
                        unchanged_data=unchanged_data,
                        unchanged_desc_data=unchanged_desc_data,
                    )
                    # add intensity targets progress to df
                    targets_int_phys_progress_full = (
                        await process_dataframe_data(
                            data=targets_int_progress_data.get(
                                "int_ph_progress_intensity_target", {}
                            ),
                            desc=targets_int_progress_desc.get(
                                "int_ph_progress_intensity_target", {}
                            ),
                            df_full=targets_int_phys_progress_full,
                            unchanged_data=unchanged_data,
                            unchanged_desc_data=unchanged_desc_data,
                        )
                    )
                    # add targets intensity economic
                    targets_int_econ_full = await process_dataframe_data(
                        data=targets_int_data.get("target_int_economic", {}),
                        desc=targets_int_desc.get("target_int_economic", {}),
                        df_full=targets_int_econ_full,
                        unchanged_data=unchanged_data,
                        unchanged_desc_data=unchanged_desc_data,
                    )
                    # add intensity targets progress to df
                    targets_int_econ_progress_full = (
                        await process_dataframe_data(
                            data=targets_int_progress_data.get(
                                "int_ec_progress_intensity_economic_target",
                                {},
                            ),
                            desc=targets_int_progress_desc.get(
                                "int_ec_progress_intensity_economic_target",
                                {},
                            ),
                            df_full=targets_int_econ_progress_full,
                            unchanged_data=unchanged_data,
                            unchanged_desc_data=unchanged_desc_data,
                        )
                    )
            if not any(
                key_ov in fe_aum_overview_headers_comparison for key_ov in res
            ):
                fe_df_mapping.pop("aum_overview", None)

            if not any(
                key_ov in fe_gross_exp_overview_headers_comparison
                for key_ov in res
            ):
                fe_df_mapping.pop("ge_overview", None)

            res_set = set(res)

            key_to_mapping = {
                "fn_aum_coverage": "aum_abs",
                "fn_gross_exp_coverage": "ge_abs",
                "fn_aum_intensity": "aum_int",
                "fn_gross_exp_intensity": "ge_int",
                "fn_aum_data_quality": "aum_dq",
                "fn_gross_exp_data_quality": "ge_dq",
            }
            if org.company_type == "Financial":
                for key_fe, mapping in key_to_mapping.items():
                    if key_fe not in res_set:
                        fe_df_mapping.pop(mapping, None)
                if fe_df_mapping:
                    for key, value in fe_df_mapping.items():
                        (
                            full_df,
                            column_key,
                            res_key,
                            namespace,
                            fe_headers,
                        ) = (
                            value["df"],
                            value["column_key"],
                            value["res_key"],
                            value["namespace"],
                            value["headers"],
                        )
                        if res_key:
                            try:
                                if res[res_key] not in [
                                    None,
                                ]:
                                    for entry in res[res_key]:
                                        entry_data = {
                                            **scope_root_data,
                                            **entry,
                                        }
                                        (
                                            processed_data,
                                            processed_description,
                                        ) = await process_financed_emissions(
                                            d=entry_data,
                                            session=self.session,
                                            source=source_value,
                                            data_source_list=DATA_SOURCE_LIST,
                                            static_cache=self.static_cache,
                                        )
                                        processed_description[
                                            "legal_entity_identifier"
                                        ] = "Company LEI"
                                        processed_description[
                                            "company_name"
                                        ] = "Company Name"
                                        if not fe_headers:
                                            temp_headers_df = pd.DataFrame(
                                                index=[0, 1],
                                                columns=list(
                                                    processed_description.keys()
                                                ),
                                            )
                                            temp_headers_df.iloc[0] = [
                                                key.upper()
                                                for key in processed_description.keys()
                                            ]
                                            temp_headers_df.iloc[1] = list(
                                                processed_description.values()
                                            )
                                            full_df = pd.concat(
                                                [
                                                    full_df,
                                                    temp_headers_df,
                                                ],
                                                ignore_index=True,
                                            )
                                            fe_df_mapping[key]["headers"] = (
                                                True
                                            )
                                        else:
                                            new_headers = [
                                                key
                                                for key in processed_description.keys()
                                                if key not in full_df.columns
                                            ]
                                            if new_headers:
                                                new_headers_df = pd.DataFrame(
                                                    index=[0, 1],
                                                    columns=list(new_headers),
                                                )
                                                new_headers_df.iloc[0] = [
                                                    key.upper()
                                                    for key in new_headers
                                                ]
                                                new_headers_df.iloc[1] = [
                                                    processed_description[key]
                                                    for key in new_headers
                                                ]
                                                full_df = pd.concat(
                                                    [
                                                        full_df,
                                                        new_headers_df,
                                                    ],
                                                    axis=1,
                                                )
                                        temp_df = pd.DataFrame(
                                            [list(processed_data.values())],
                                            columns=list(
                                                processed_description.keys()
                                            ),
                                        )
                                        full_df = pd.concat(
                                            [full_df, temp_df],
                                            ignore_index=True,
                                        )
                            except KeyError:
                                full_df = pd.DataFrame(
                                    columns=namespace[column_key]
                                )
                        else:
                            (
                                extracted_data,
                                extracted_desc,
                            ) = await extract_data_from_fe(
                                d=res,
                                keys=namespace[column_key],
                                session=self.session,
                                source=source_value,
                                last_updated=last_updated,
                                data_source_list=DATA_SOURCE_LIST,
                                static_cache=self.static_cache,
                            )
                            fi_mappings = {
                                "legal_entity_identifier": "Company LEI",
                                "company_name": "Company Name",
                            }
                            for k, d in fi_mappings.items():
                                if k in extracted_data:
                                    extracted_desc[k] = d
                            if not fe_headers:
                                temp_headers_df = pd.DataFrame(
                                    index=[0, 1],
                                    columns=list(extracted_desc.keys()),
                                )
                                temp_headers_df.iloc[0] = [
                                    key.upper()
                                    for key in extracted_desc.keys()
                                ]
                                temp_headers_df.iloc[1] = list(
                                    extracted_desc.values()
                                )
                                full_df = pd.concat(
                                    [full_df, temp_headers_df],
                                    ignore_index=True,
                                )
                                fe_df_mapping[key]["headers"] = True
                            else:
                                new_headers = [
                                    key
                                    for key in extracted_desc.keys()
                                    if key not in full_df.columns
                                ]
                                if new_headers:
                                    new_headers_df = pd.DataFrame(
                                        index=[0, 1],
                                        columns=list(new_headers),
                                    )
                                    new_headers_df.iloc[0] = [
                                        key.upper() for key in new_headers
                                    ]
                                    new_headers_df.iloc[1] = [
                                        extracted_desc[key]
                                        for key in new_headers
                                    ]
                                    full_df = pd.concat(
                                        [full_df, new_headers_df],
                                        axis=1,
                                    )
                            temp_df = pd.DataFrame(
                                [list(extracted_data.values())],
                                columns=list(extracted_desc.keys()),
                            )
                            full_df = pd.concat(
                                [full_df, temp_df], ignore_index=True
                            )
                            full_df = full_df[temp_df.columns]

                        fe_df_mapping[key]["df"] = full_df
                if fe_df_mapping.get("aum_overview", None):
                    fe_aum_df = fe_df_mapping["aum_overview"]["df"]
                if fe_df_mapping.get("ge_overview", None):
                    fe_ge_df = fe_df_mapping["ge_overview"]["df"]
                if fe_df_mapping.get("aum_abs", None):
                    fn_aum_abs_full_df = fe_df_mapping["aum_abs"]["df"]
                if fe_df_mapping.get("ge_abs", None):
                    fn_ge_abs_full_df = fe_df_mapping["ge_abs"]["df"]
                if fe_df_mapping.get("aum_int", None):
                    fn_aum_int_full_df = fe_df_mapping["aum_int"]["df"]
                if fe_df_mapping.get("ge_int", None):
                    fn_ge_int_full_df = fe_df_mapping["ge_int"]["df"]
                if fe_df_mapping.get("aum_dq", None):
                    fn_aum_dq_full_df = fe_df_mapping["aum_dq"]["df"]
                if fe_df_mapping.get("ge_dq", None):
                    fn_ge_dq_full_df = fe_df_mapping["ge_dq"]["df"]
        if scope_3_methodology_full.empty:
            # 3 methodology
            scope_3_methodology_full = pd.DataFrame(
                index=[0, 1],
                columns=list(scope_3_methodology_headers.keys()),
            )
            scope_3_methodology_full.iloc[0] = [
                key.upper() for key in scope_3_methodology_headers.keys()
            ]
            scope_3_methodology_full.iloc[1] = list(
                scope_3_methodology_headers.values()
            )
        else:
            scope_3_methodology_full = pd.concat(
                [
                    scope_3_methodology_full,
                    *[
                        methodology_ghgp_dataframes[
                            f"scope_3_methodology_c{i}_full"
                        ]
                        for i in range(1, 16)
                    ],
                    methodology_ghgp_dataframes[
                        "scope_3_methodology_other_full"
                    ],
                    *[
                        methodology_iso_dataframes[
                            f"scope_3_methodology_c{i}_full"
                        ]
                        for i in range(1, 16)
                    ],
                    methodology_iso_dataframes[
                        "scope_3_methodology_other_full"
                    ],
                ],
                axis=1,
                ignore_index=True,
            )
        if scope_3_exclusion_full.empty:
            # 3 exclusion
            scope_3_exclusion_full = pd.DataFrame(
                index=[0, 1],
                columns=list(scope_3_exclusion_headers.keys()),
            )
            scope_3_exclusion_full.iloc[0] = [
                key.upper() for key in scope_3_exclusion_headers.keys()
            ]
            scope_3_exclusion_full.iloc[1] = list(
                scope_3_exclusion_headers.values()
            )
        else:
            scope_3_exclusion_full = pd.concat(
                [
                    scope_3_methodology_full,
                    *[
                        exclusion_ghgp_dataframes[
                            f"scope_3_exclusion_c{i}_full"
                        ]
                        for i in range(1, 16)
                    ],
                    exclusion_ghgp_dataframes["scope_3_exclusion_other_full"],
                    *[
                        exclusion_iso_dataframes[
                            f"scope_3_exclusion_c{i}_full"
                        ]
                        for i in range(1, 16)
                    ],
                    exclusion_iso_dataframes["scope_3_exclusion_other_full"],
                ],
                axis=1,
                ignore_index=True,
            )
        # dataframe mapping
        fe_sheet_and_df_mapping = [
            (SearchSheets.KEY.value, key_full, True),
            (SearchSheets.METADATA.value, metadata_full, False),
            (
                SearchSheets.SCOPE_1_EMISSIONS_SHEET.value,
                scope_emissions_full,
                False,
            ),
            (
                SearchSheets.SCOPE_1_GHG_BREAKDOWN_SHEET.value,
                scope_ghg_full,
                False,
            ),
            (
                SearchSheets.SCOPE_1_EXCLUSIONS_SHEET.value,
                scope_exclusion_full,
                False,
            ),
            (
                SearchSheets.SCOPE_2_LB_EMISSIONS_SHEET.value,
                scope_lb_emissions_full,
                False,
            ),
            (
                SearchSheets.SCOPE_2_LB_EXCLUSIONS_SHEET.value,
                scope_lb_exclusion_full,
                False,
            ),
            (
                SearchSheets.SCOPE_2_MB_EMISSIONS_SHEET.value,
                scope_mb_emissions_full,
                False,
            ),
            (
                SearchSheets.SCOPE_2_MB_EXCLUSIONS_SHEET.value,
                scope_mb_exclusion_full,
                False,
            ),
            (
                SearchSheets.SCOPE_3_EMISSIONS_SHEET.value,
                scope_3_emissions_full,
                False,
            ),
            (
                SearchSheets.SCOPE_3_METHODOLOGY_SHEET.value,
                scope_3_methodology_full,
                False,
            ),
            (
                SearchSheets.SCOPE_3_EXCLUSIONS_SHEET.value,
                scope_3_exclusion_full,
                False,
            ),
            (SearchSheets.FE_OVERVIEW_AUM_SHEET.value, fe_aum_df, False),
            (
                SearchSheets.FE_ABSOLUTE_AUM_SHEET.value,
                fn_aum_abs_full_df,
                False,
            ),
            (
                SearchSheets.FE_INTENSITY_AUM_SHEET.value,
                fn_aum_int_full_df,
                False,
            ),
            (
                SearchSheets.FE_DATA_QUALITY_AUM_SHEET.value,
                fn_aum_dq_full_df,
                False,
            ),
            (SearchSheets.FE_OVERVIEW_GE_SHEET.value, fe_ge_df, False),
            (
                SearchSheets.FE_ABSOLUTE_GE_SHEET.value,
                fn_ge_abs_full_df,
                False,
            ),
            (
                SearchSheets.FE_INTENSITY_GE_SHEET.value,
                fn_ge_int_full_df,
                False,
            ),
            (
                SearchSheets.FE_DATA_QUALITY_GE_SHEET.value,
                fn_ge_dq_full_df,
                False,
            ),
            (SearchSheets.AV_EMISSIONS_SHEET.value, assure_verif_full, False),
            (
                SearchSheets.RESTATEMENTS_EMISSIONS_SHEET.value,
                restatements_e_df,
                False,
            ),
            (
                SearchSheets.TARGETS_ABSOLUTE_SHEET.value,
                targets_abs_full,
                False,
            ),
            (
                SearchSheets.TARGETS_PROGRESS_ABSOLUTE_SHEET.value,
                targets_abs_progress_full,
                False,
            ),
            (
                SearchSheets.TARGETS_INTENSITY_PHYS_SHEET.value,
                targets_int_phys_full,
                False,
            ),
            (
                SearchSheets.TARGETS_INTENSITY_ECON_SHEET.value,
                targets_int_econ_full,
                False,
            ),
            (
                SearchSheets.TARGETS_PROGRESS_PHYS_INTENSITY_SHEET.value,
                targets_int_phys_progress_full,
                False,
            ),
            (
                SearchSheets.TARGETS_PROGRESS_ECON_INTENSITY_SHEET.value,
                targets_int_econ_progress_full,
                False,
            ),
            (
                SearchSheets.AV_TARGETS_SHEET.value,
                assure_verif_targets_full,
                False,
            ),
            (
                SearchSheets.RESTATEMENTS_TARGETS_SHEET.value,
                restatements_t_df,
                False,
            ),
        ]
        skip_sheets = [
            SearchSheets.KEY.value,
            SearchSheets.METADATA.value,
            SearchSheets.TARGETS_ABSOLUTE_SHEET.value,
            SearchSheets.TARGETS_PROGRESS_ABSOLUTE_SHEET.value,
            SearchSheets.TARGETS_INTENSITY_PHYS_SHEET.value,
            SearchSheets.TARGETS_INTENSITY_ECON_SHEET.value,
            SearchSheets.TARGETS_PROGRESS_PHYS_INTENSITY_SHEET.value,
            SearchSheets.TARGETS_PROGRESS_ECON_INTENSITY_SHEET.value,
            SearchSheets.AV_TARGETS_SHEET,
        ]
        restatements_sheets = [
            SearchSheets.RESTATEMENTS_EMISSIONS_SHEET.value,
            SearchSheets.RESTATEMENTS_TARGETS_SHEET.value,
        ]
        # pylint: disable=abstract-class-instantiated
        with ExcelWriter(excel_filename) as writer:  # pylint: disable=abstract-class-instantiated
            # add mapped dataframes to excel
            for sheet_name, df, header in fe_sheet_and_df_mapping:
                if sheet_name not in skip_sheets:
                    if df.iloc[2:].empty:
                        data_to_copy = metadata_full.iloc[2:, [0, 1, 6, 7, 10]]
                        if "FIN" in sheet_name:
                            data_to_copy = data_to_copy[
                                data_to_copy.iloc[:, 0].isin(financial_lei)
                            ]
                        df = pd.concat(
                            [
                                df.iloc[:2],
                                data_to_copy.reset_index(drop=True),
                            ],
                            ignore_index=True,
                        )
                if source:
                    # check if we download-all source or timestamp file
                    # skip restatements worksheets
                    if sheet_name in restatements_sheets:
                        continue
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=header,
                    float_format="%.2f",
                )
                ws = writer.sheets[sheet_name]
                if not header:
                    for cell in ws[2]:
                        cell.border = thick_border_bottom
                else:
                    normal_font = Font(bold=False)
                    for cell in ws[1]:
                        cell.font = normal_font
                        cell.border = normal_border
        return excel_filename

    async def _generate_excel_with_search_query_fields(
        self, filename: str | None = None
    ) -> str:
        """
        Generate excel worksheets from data frames with search query fields

        Returns:
            Excel filename for downloading
        """
        excel_filename = "Data_Explorer.xlsx" if filename is None else filename
        thick_border_bottom = Border(bottom=Side(style="thin"))
        # load formatted attributes json

        submission_loader = SubmissionLoader(
            redis_cache=self.cache,
            session=self.session,
            core_cache=self.static_cache,
        )
        await self._process_attributes_fields_and_values(
            submission_loader=submission_loader
        )
        # create default full
        default_full = pd.DataFrame(
            index=[0, 1], columns=list(self.default_single_headers.keys())
        )
        default_full.iloc[0] = [
            key.upper() for key in self.default_single_headers.keys()
        ]
        default_full.iloc[1] = [
            transform_subscript_to_normal(x)
            for x in self.default_single_headers.values()
        ]
        # fill values for every row
        for single_value in self.attributes_single_values:
            data_df = pd.DataFrame(
                [list(single_value.values())],
                columns=list(single_value.keys()),
            )
            default_full = pd.concat(
                [default_full, data_df], ignore_index=True
            )
        # pylint: disable=abstract-class-instantiated
        with ExcelWriter(excel_filename) as writer:  # pylint: disable=abstract-class-instantiated
            default_full.to_excel(
                writer, sheet_name="MAIN", index=False, header=False
            )
            ws = writer.sheets["MAIN"]
            for cell in ws[2]:
                cell.border = thick_border_bottom
        return excel_filename

    async def download_excel(
        self,
        filename: str | None = None,
        source: bool = False,
        last_updated: bool = False,
        down_all: bool = False,
    ) -> str:
        """
        Main method of this class.

        It constructs download to excel with query_result dict

        Returns:
            list: The Excel filename.
        """
        if down_all:
            if not source:
                submission_loader = SubmissionLoader(
                    redis_cache=self.cache,
                    session=self.session,
                    core_cache=self.static_cache,
                )
                await self._process_submission_rest(submission_loader)
            logger.info("Processed submission restatements")
            excel_filename = await self._generate_excel_for_all(
                filename=filename, source=source, last_updated=last_updated
            )
        else:
            excel_filename = (
                await self._generate_excel_with_search_query_fields(
                    filename=filename
                )
            )
        return excel_filename

    async def classification_toggler(self) -> bool:
        """
        this method return config value for classification

        Returns:
            bool: of classification
        """
        # Create a select query
        stmt = select(Config.value).where(
            Config.name == "data_download.exclude_classification"
        )

        # Execute the query
        result = await self.session.execute(stmt)

        # Fetch the scalar result
        sics_sector_value = result.scalar()

        return sics_sector_value
