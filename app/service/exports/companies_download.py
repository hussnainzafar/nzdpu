from dataclasses import dataclass, field

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import Executable
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import (
    Organization,
)
from app.db.redis import RedisClient
from app.db.types import EN_DASH, NullTypeState
from app.routers.utils import (
    get_choice_value,
    scientific_to_float,
)
from app.schemas.companies import CompanyEmissions, HistoryItem
from app.service.core.cache import CoreMemoryCache
from app.service.core.loaders import SubmissionLoader
from app.service.exports.forms_processor import (
    process_assure_verif_companies,
    process_company_metadata,
    process_data_export,
    process_scope_emissions,
    process_target_validation_companies,
    process_targets_export,
)
from app.service.exports.headers.headers import (
    CompaniesSheets,
    ExportOptions,
    get_companies_headers,
)
from app.service.exports.restatement import RestatementExportManager
from app.service.exports.utils import (
    combine_units_into_one_list,
    financed_emissions_formatter,
    format_datetime_for_downloads,
    get_constraint_views,
    remove_fields_from_form,
    return_unit_from_field,
    scope_emissions_formatter,
    update_dataframe,
)
from app.service.history_service import HistoryService
from app.service.schema_service import FormGroupBy
from app.service.utils import (
    parse_and_transform_subscripts_to_normal,
    transform_subscript_to_normal,
)
from app.utils import excel_filename_sics, sanitize_filename


@dataclass
class HeaderConfig:
    """
    HeaderConfig class provides configuration mappings for different headers.
    """

    @staticmethod
    def _headers_metadata_config():
        """
        Returns configuration for company metadata headers.
        """
        metadata = get_companies_headers(
            CompaniesSheets.company_metadata.value
        )
        return {
            "company_metadata_full": metadata,
            "metadata_fields_and_desc_full": metadata,
        }

    @staticmethod
    def _headers_scope_config() -> dict:
        """
        Returns configuration for scope-related headers.
        """
        default = {
            "scope_emissions_full": "SCOPE_1_EMISSIONS",
            "scope_exclusion_full": "SCOPE_1_EXCLUSION",
            "scope_emissions_change_full": "SCOPE_1_EMISSIONS_CHANGE",
            "scope_ghg_full": "SCOPE_1_GHG_BREAKDOWN",
            "scope_ghg_other_full": "SCOPE_1_GHG_BREAKDOWN_OTHER",
            "scope_lb_emissions_full": "SCOPE_2_LB_EMISSIONS",
            "scope_lb_exclusion_full": "SCOPE_2_LB_EXCLUSION",
            "scope_lb_change_full": "SCOPE_2_LB_CHANGE",
            "scope_mb_emissions_full": "SCOPE_2_MB_EMISSIONS",
            "scope_mb_exclusion_full": "SCOPE_2_MB_EXCLUSION",
            "scope_mb_change_full": "SCOPE_2_MB_CHANGE",
            "scope_3_ghgp_emissions_full": "SCOPE_3_GHGP",
        }
        for i in range(1, 16):
            scope_3_categories = {
                f"scope_3_ghgp_c{i}_full": f"SCOPE_3_GHGP_C{i}",
                f"scope_3_ghgp_c{i}_method_full": f"SCOPE_3_GHGP_C{i}_METHODOLOGY",
                f"scope_3_ghgp_c{i}_disclose_full": f"SCOPE_3_GHGP_C{i}_EXCLUSION_DISCLOSE",
                f"scope_3_ghgp_c{i}_exclusion_full": f"SCOPE_3_GHGP_C{i}_EXCLUSION",
                f"scope_3_ghgp_c{i}_change_full": f"SCOPE_3_GHGP_C{i}_CHANGE",
            }
            default.update(scope_3_categories)
        default.update(
            {
                "scope_3_ghgp_other_full": "SCOPE_3_GHGP_OTHER",
                "scope_3_ghgp_other_method_full": "SCOPE_3_GHGP_OTHER_METHODOLOGY",
                "scope_3_ghgp_other_disclose_full": "SCOPE_3_GHGP_OTHER_EXCLUSION_DISCLOSE",
                "scope_3_ghgp_other_exclusion_full": "SCOPE_3_GHGP_OTHER_EXCLUSION",
                "scope_3_ghgp_other_change_full": "SCOPE_3_GHGP_OTHER_CHANGE",
            }
        )
        for i in range(3, 7):
            scope_3_iso_categories = {
                f"scope_3_iso_c{i}_full": f"SCOPE_3_ISO_C{i}",
                f"scope_3_iso_c{i}_method_full": f"SCOPE_3_ISO_C{i}_METHODOLOGY",
                f"scope_3_iso_c{i}_disclose_full": f"SCOPE_3_ISO_C{i}_EXCLUSION_DISCLOSE",
                f"scope_3_iso_c{i}_exclusion_full": f"SCOPE_3_ISO_C{i}_EXCLUSION",
                f"scope_3_iso_c{i}_change_full": f"SCOPE_3_ISO_C{i}_CHANGE",
            }
            default.update(scope_3_iso_categories)

        return default

    @staticmethod
    def _headers_scope_categories_config(idx: int) -> dict:
        """
        Returns a mapping of suffixes to headers for a given category index `idx`.
        """
        suffixes = ["", "_METHODOLOGY", "_EXCLUSION", "_CHANGE"]
        scope_categories = {}

        base_key = (
            f"SCOPE_3_GHGP_C{idx}" if idx != "OTHER" else "SCOPE_3_GHGP_OTHER"
        )

        for suffix in suffixes:
            key_name = f"scope_3_ghgp_category{suffix.lower()}_full"
            scope_categories[key_name] = f"{base_key}{suffix}"

        return scope_categories

    @staticmethod
    def _headers_financed_config() -> dict:
        """
        Returns configuration for financed-related headers.
        """
        return {
            "fe_aum_overview_full": "FE_AUM_OVERVIEW",
            "fe_aum_absolute_full": "FE_AUM_ABS_COVERAGE",
            "fe_aum_change_full": "FE_AUM_CHANGE",
            "fe_aum_intensity_full": "FE_AUM_INTENSITY",
            "fe_aum_data_quality_full": "FE_AUM_DATA_QUALITY",
            "fe_ge_overview_full": "FE_GE_OVERVIEW",
            "fe_ge_absolute_full": "FE_GE_AUM_COVERAGE",
            "fe_ge_change_full": "FE_GE_CHANGE",
            "fe_ge_intensity_full": "FE_GE_INTENSITY",
            "fe_ge_data_quality_full": "FE_GE_DATA_QUALITY",
        }

    @staticmethod
    def _headers_assure_verif_config() -> dict:
        """
        Returns configuration for assurance and verification headers.
        """
        return {
            "rationale_verif_emissions_non_disclose": "Rationale if verification of GHG emissions not disclosed",
            "verif_emissions_provider_1": "Assurance or verification provider (1)",
            "verif_emissions_standard_1": "Standard emissions assured or verified against (1)",
            "verif_emissions_level_of_assurance_1": "Level of assurance provided for emissions (1)",
            "verif_emissions_cvg_scope_1": "Scopes included in assurance or verification statement (1)",
            "verif_emissions_cvg_s3_cat_1": "Scope 3 categories included in assurance or verification statement (1)",
            "perc_s1_emissions_verif_1": "Percentage of total Scope 1 GHG emissions assured or verified (1)",
            "perc_s2_lb_emissions_verif_1": "Percentage of total Scope 2 location-based GHG emissions assured or verified (1)",
            "perc_s2_mb_emissions_verif_1": "Percentage of total Scope 2 market-based GHG emissions assured or verified (1)",
            "perc_s3_emissions_verif_1": "Percentage of total Scope 3 GHG emissions assured or verified (1)",
            "verif_emissions_statement_1": "Assurance or verification of GHG emissions statement (1)",
        }

    @staticmethod
    def _no_data_default_headers(
        worksheet: CompaniesSheets, plural: bool = False
    ) -> list:
        df_list = []
        targets = "Targets" if plural else "Target"
        default_value = f"No Emissions Reduction {targets} Data Available."
        data_row = {
            "Disclosure Year": default_value,
            "Target ID": None,
            "Target Name": None,
            "Field Name": None,
            "Short Description": None,
            "Units": None,
            "Reporting Year": None,
            "Value": None,
            "Source": None,
            "Last Updated": None,
        }
        pop_headers_worksheet = {
            CompaniesSheets.emissions_reduction_targets: ["Reporting Year"],
            CompaniesSheets.validation_sheet: ["Units", "Reporting Year"],
        }
        for key in pop_headers_worksheet.get(worksheet, []):
            data_row.pop(key, None)
        df_list.append(data_row)
        return df_list


@dataclass
class CompaniesExportManager(HeaderConfig):
    """
    CompaniesExportManager class.
    """

    session: AsyncSession
    static_cache: CoreMemoryCache
    cache: RedisClient
    result: CompanyEmissions
    stmt: Executable
    submission_loader: SubmissionLoader
    company: Organization
    download_option: str = ExportOptions.COMPANIES.value
    forms_group_by: list[FormGroupBy] = field(default_factory=lambda: [])
    history_service: HistoryService = HistoryService()

    @staticmethod
    def _make_default_extra_headers(
        dataframes_list: list[pd.DataFrame], restated=True
    ):
        for dataframe in dataframes_list:
            for year in range(2015, 2024):
                default = {
                    f"Value_{year}": [],
                    f"Source_{year}": [],
                    f"Last Updated_{year}": [],
                }
                # do not load restated for metadata worksheet
                if restated:
                    default.update({f"Restated_{year}": []})
                for column_name, column_data in default.items():
                    if column_name not in dataframe.columns:
                        dataframe[column_name] = pd.Series(column_data)

    @staticmethod
    def _extract_and_assign(data_dict, keys):
        extracted_data = {key: data_dict.pop(key, None) for key in keys}
        return extracted_data

    @staticmethod
    async def _handle_process_methods(handler, *args, **kwargs):
        try:
            return await handler(*args, **kwargs)
        except KeyError:
            return {}, {}, {}, {}, {}

    @staticmethod
    def _fill_no_disclosure_data(df: pd.DataFrame):
        """
        Replace all-None columns with 'blank' which represent
        no disclosure for that year
        """
        # Identify target columns
        target_columns = {
            "source_last_updated": [
                col
                for col in df.columns
                if col.startswith("Source_") or col.startswith("Last Updated_")
            ],
            "value_columns": [
                col for col in df.columns if col.startswith("Value_")
            ],
        }

        for col_group, replacement_value in [
            (target_columns["source_last_updated"], "blank"),
            (target_columns["value_columns"], NullTypeState.LONG_DASH),
        ]:
            for col in col_group:
                if df[col].isna().all():
                    df[col] = df[col].fillna(replacement_value)

    async def _make_default_headers(
        self,
        headers: dict = None,
        sheet: str = None,
        units: bool = False,
        targets: bool = False,
    ):
        if targets:
            target_dataframe = {
                "Target ID": [],
                "Target Name": [],
                "Field Name": [],
                "Short Description": [],
                "Units": [],
            }
            return pd.DataFrame(target_dataframe)
        field_names = list(headers.keys())
        descriptions = list(headers.values())

        dataframe_full = {
            "Field Name": field_names,
            "Short Description": descriptions,
        }

        if units:
            dataframe_full["Units"] = [
                (
                    transform_subscript_to_normal(
                        await return_unit_from_field(
                            key, self.session, self.static_cache, sheet
                        )
                    )
                )
                for key in field_names
            ]

        return pd.DataFrame(dataframe_full)

    async def get_restatements_and_units(self):
        submission_loader = SubmissionLoader(
            self.session, self.static_cache, self.cache
        )
        restated_submissions = [
            h.submission
            for h in self.result.history
            if h.submission.restated_fields_data_source
        ]
        if restated_submissions:
            restatements_list = []
            for active_submission in restated_submissions:
                restatements_manager = RestatementExportManager(
                    submissions=[active_submission.model_dump()],
                    session=self.session,
                    cache=self.cache,
                    static_cache=self.static_cache,
                    extract_targets=True,
                )
                restatements = (
                    await restatements_manager.get_restatements_nz_id_mapping(
                        [active_submission.id]
                    )
                )
                if restatements:
                    restated_attrs_path_lookup = (
                        restatements_manager.restated_attr_paths_mapping(
                            restatements
                        )
                    )
                    original_submissions = (
                        await restatements_manager.get_original_submissions(
                            restated_attrs_path_lookup, restatements
                        )
                    )
                    company_restatements = restatements[self.company.nz_id]
                    attr_paths = restated_attrs_path_lookup[self.company.nz_id]
                    constraint_views = await get_constraint_views(
                        self.session, attr_paths
                    )
                    for field_name, field_value in attr_paths.items():
                        await restatements_manager.extract_restated_field(
                            self.company.nz_id,
                            field_name,
                            field_value,
                            original_submissions,
                            active_submission.model_dump(),
                            company_restatements,
                            constraint_views,
                            submission_loader,
                        )
                    combine_rest = (
                        restatements_manager.restatements_emissions
                        + restatements_manager.restatements_targets
                    )
                    restatements_list.extend(combine_rest)
            return restatements_list

    def _handle_grouping(self, context=None):
        """
        Modify the structure for grouping only when required.
        Skip grouping when the context is 'excel_download'.
        """
        if context == "excel_download":
            # Skip grouping for Excel generation
            return

        # Perform grouping logic
        self.result.history = [
            {
                "reporting_year": x.reporting_year,
                "submission": x.submission,
            }
            for x in self.result.history
        ]

        self.history_service.group_form_items(
            self.forms_group_by, self.result.history
        )

        # modify the structure back how it was to fit the generate file function
        self.result.history = [
            HistoryItem(
                reporting_year=x.get("reporting_year"),
                submission=x.get("submission"),
            )
            for x in self.result.history
        ]

    async def generate_companies_download(
        self, exclude_classification_forced: bool | None = None
    ):
        # get source ids for mapping restated and source columns
        choices = await self.static_cache.choices()
        source_set = list(
            filter(lambda x: x.set_name == "source_list", choices.values())
        )
        try:
            restatement_list = await self.get_restatements_and_units()
        except Exception:
            restatement_list = []
        # add restatements data frame
        restatements_df = pd.DataFrame(restatement_list)
        extension = ".xlsx"
        excel_filename = (
            sanitize_filename(
                prefix="nzdpu",
                company=self.company.legal_name,
                nz_id=str(self.company.nz_id),
            )
            + excel_filename_sics(exclude_classification_forced)
            + extension
        )
        # default headers
        main_headers = get_companies_headers(CompaniesSheets.main_sheet.value)
        financed_emissions_headers = get_companies_headers(
            CompaniesSheets.financed_emissions_sheet.value
        )
        assure_verif_headers = get_companies_headers(
            CompaniesSheets.assure_verif.value
        )
        metadata_results = {}
        for key, header_key in self._headers_metadata_config().items():
            metadata_results[key] = await self._make_default_headers(
                headers=header_key
            )
        company_metadata_full = metadata_results["company_metadata_full"]
        metadata_fields_and_desc_full = metadata_results[
            "metadata_fields_and_desc_full"
        ]

        scope_results = {}
        for key, header_key in self._headers_scope_config().items():
            scope_results[key] = await self._make_default_headers(
                headers=main_headers.get(header_key),
                sheet="scope",
                units=True,
            )
        # format dataframes for all scope emissions
        dataframes_scope_results = {
            scope: scope_results[scope] for scope in scope_results
        }
        date_fields_and_desc_full = pd.DataFrame(
            {
                "Field Name": [],
                "Short Description": [],
                "Units": [],
            }
        )

        # check if company type is Financial to process financed emissions
        if self.company.company_type == "Financial":
            financed_results = {}
            for key, header_key in self._headers_financed_config().items():
                financed_results[key] = await self._make_default_headers(
                    headers=financed_emissions_headers.get(header_key),
                    sheet="financed",
                    units=True,
                )
            fe_aum_overview_full = financed_results["fe_aum_overview_full"]
            fe_aum_absolute_full = financed_results["fe_aum_absolute_full"]
            fe_aum_change_full = financed_results["fe_aum_change_full"]
            fe_aum_intensity_full = financed_results["fe_aum_intensity_full"]
            fe_aum_data_quality_full = financed_results[
                "fe_aum_data_quality_full"
            ]
            fe_ge_overview_full = financed_results["fe_ge_overview_full"]
            fe_ge_absolute_full = financed_results["fe_ge_absolute_full"]
            fe_ge_change_full = financed_results["fe_ge_change_full"]
            fe_ge_intensity_full = financed_results["fe_ge_intensity_full"]
            fe_ge_data_quality_full = financed_results[
                "fe_ge_data_quality_full"
            ]
            fe_complete_full = pd.DataFrame(
                {
                    "Field Name": [],
                    "Short Description": [],
                    "Units": [],
                }
            )
        # create assure verification data frame
        assure_verif_temp = {}
        # create targets data frames
        targets_full_list = []
        targets_progress_full_list = []
        # create targets progress data frame
        targets_abs_progress_full = await self._make_default_headers(
            targets=True
        )
        targets_int_progress_full = await self._make_default_headers(
            targets=True
        )
        # make default extra headers for df who have restated field
        restated_headers = [
            dataframes_scope_results[key] for key in scope_results
        ] + [
            date_fields_and_desc_full,
            targets_abs_progress_full,
            targets_int_progress_full,
        ]
        self._make_default_extra_headers(restated_headers)
        self._make_default_extra_headers(
            [company_metadata_full, metadata_fields_and_desc_full],
            restated=False,
        )
        if self.company.company_type == "Financial":
            self._make_default_extra_headers(
                [
                    fe_aum_overview_full,
                    fe_aum_absolute_full,
                    fe_aum_change_full,
                    fe_aum_intensity_full,
                    fe_aum_data_quality_full,
                    fe_ge_overview_full,
                    fe_ge_absolute_full,
                    fe_ge_change_full,
                    fe_ge_intensity_full,
                    fe_ge_data_quality_full,
                    fe_complete_full,
                ]
            )
        # create assure verification validation data frame
        validation_full_list = []
        reported_year_list = []
        last_updated_map = {}

        # skip the grouping for excel download to not break the data consistency
        self._handle_grouping(context="excel_download")

        # this is used to compare all the fields in the emissions so that we can keep the longest sub form data
        data_fields_emissions_tab = [pd.DataFrame() for _ in range(112)]

        for res_new in self.result.history:
            # load submission values with null type 3 rendering
            res_values = parse_and_transform_subscripts_to_normal(
                res_new.submission.values
            )
            # load units from submissions
            res_units = parse_and_transform_subscripts_to_normal(
                res_new.submission.units
            )
            # load reporting year
            reporting_year = res_new.reporting_year
            # load disclosure year
            disclosure_year = res_values.get("filing_year")
            # add reporting year to list for loading last_updated where we
            # don't have data
            reported_year_list.append(reporting_year)
            source = res_values.get("disclosure_source")
            # get source choice value
            source = (
                await get_choice_value(source, self.session, self.static_cache)
                if isinstance(source, int)
                else source
            )
            last_updated = format_datetime_for_downloads(
                res_values.get(
                    "reporting_datetime", res_new.submission.created_on
                )
            )
            # add last_updated map for adding empty cell in last updated col
            last_updated_map[reporting_year] = last_updated
            # get restated dict to check if this year is restated
            restated = (
                True
                if res_new.submission.restated_fields_data_source
                else False
            )
            restated_fields = parse_and_transform_subscripts_to_normal(
                res_new.submission.restated_fields_data_source
            )
            # process company metadata
            (
                company_metadata_data,
                company_metadata_desc,
            ) = await process_company_metadata(
                d=res_values,
                company=self.company,
                reporting_year=reporting_year,
                session=self.session,
                static_cache=self.static_cache,
                exclude_classification_forced=exclude_classification_forced,
            )
            # process scope 1 emissions
            scope_1_emissions_fields = scope_emissions_formatter(
                res_values, "s1_emissions"
            )
            (
                scope_emissions_data,
                scope_emissions_desc,
                scope_emissions_source,
                scope_emissions_reported,
                scope_emissions_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_1_emissions_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 1 emissions change type
            scope_1_change_types_fields = scope_emissions_formatter(
                res_values, "s1_change_type"
            )
            (
                scope_change_type_data,
                scope_change_type_desc,
                scope_change_type_source,
                scope_change_type_reported,
                scope_change_type_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_1_change_types_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 1 exclusion
            (
                scope_exclusion_data,
                scope_exclusion_desc,
                scope_exclusion_source,
                scope_exclusion_reported,
                scope_exclusion_units,
            ) = await self._handle_process_methods(
                process_data_export,
                data_export_list=res_values.get("s1_emissions_exclusion_dict"),
                units_list=res_units.get("s1_emissions_exclusion_dict"),
                source=source,
                last_updated=last_updated,
                form_name="s1_emissions_exclusion_dict",
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 1 ghg breakdown
            scope_1_ghg_fields = scope_emissions_formatter(
                res_values, "s1_ghg"
            )
            (
                scope_ghg_data,
                scope_ghg_desc,
                scope_ghg_source,
                scope_ghg_reported,
                scope_ghg_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_1_ghg_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 1 ghg breakdown other
            (
                scope_ghg_other_data,
                scope_ghg_other_desc,
                scope_ghg_other_source,
                scope_ghg_other_reported,
                scope_ghg_other_units,
            ) = await self._handle_process_methods(
                process_data_export,
                data_export_list=res_values.get("s1_other_ghg_emissions_dict"),
                units_list=res_units.get("s1_other_ghg_emissions_dict"),
                source=source,
                last_updated=last_updated,
                form_name="s1_other_ghg_emissions_dict",
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 2 lb emissions
            scope_2_lb_emissions_fields = scope_emissions_formatter(
                res_values, "s2_lb_emissions"
            )
            (
                scope_lb_emissions_data,
                scope_lb_emissions_desc,
                scope_lb_emissions_source,
                scope_lb_emissions_reported,
                scope_lb_emissions_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_2_lb_emissions_fields,
                static_cache=self.static_cache,
                session=self.session,
                download_option=self.download_option,
                restated=restated_fields,
            )
            # process scope 2 lb exclusion
            (
                scope_lb_exclusion_data,
                scope_lb_exclusion_desc,
                scope_lb_exclusion_source,
                scope_lb_exclusion_reported,
                scope_lb_exclusion_units,
            ) = await self._handle_process_methods(
                process_data_export,
                data_export_list=res_values.get(
                    "s2_lb_emissions_exclusion_dict"
                ),
                units_list=res_units.get("s2_lb_emissions_exclusion_dict"),
                source=source,
                last_updated=last_updated,
                static_cache=self.static_cache,
                form_name="s2_lb_emissions_exclusion_dict",
                session=self.session,
                download_option=self.download_option,
                restated=restated_fields,
            )
            # process scope 2 mb emissions change type
            scope_2_lb_change_types_fields = scope_emissions_formatter(
                res_values, "s2_lb_change_type"
            )
            (
                scope_lb_change_type_data,
                scope_lb_change_type_desc,
                scope_lb_change_type_source,
                scope_lb_change_type_reported,
                scope_lb_change_type_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_2_lb_change_types_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 2 mb emissions
            scope_2_mb_emissions_fields = scope_emissions_formatter(
                res_values, "s2_mb_emissions"
            )
            (
                scope_mb_emissions_data,
                scope_mb_emissions_desc,
                scope_mb_emissions_source,
                scope_mb_emissions_reported,
                scope_mb_emissions_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                static_cache=self.static_cache,
                last_updated=last_updated,
                emissions=scope_2_mb_emissions_fields,
                session=self.session,
                download_option=self.download_option,
                restated=restated_fields,
            )
            # process scope 2 mb exclusion
            (
                scope_mb_exclusion_data,
                scope_mb_exclusion_desc,
                scope_mb_exclusion_source,
                scope_mb_exclusion_reported,
                scope_mb_exclusion_units,
            ) = await self._handle_process_methods(
                process_data_export,
                data_export_list=res_values.get(
                    "s2_mb_emissions_exclusion_dict"
                ),
                units_list=res_units.get("s2_mb_emissions_exclusion_dict"),
                source=source,
                last_updated=last_updated,
                form_name="s2_mb_emissions_exclusion_dict",
                session=self.session,
                static_cache=self.static_cache,
                download_option=self.download_option,
                restated=restated_fields,
            )
            # process scope 2 mb emissions change type
            scope_2_mb_change_types_fields = scope_emissions_formatter(
                res_values, "s2_mb_change_type"
            )
            (
                scope_mb_change_type_data,
                scope_mb_change_type_desc,
                scope_mb_change_type_source,
                scope_mb_change_type_reported,
                scope_mb_change_type_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_2_mb_change_types_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            # process scope 3 ghgp emissions
            scope_3_fields = scope_emissions_formatter(
                res_values, "s3_emissions"
            )
            (
                scope_3_ghgp_emissions_data,
                scope_3_ghgp_emissions_desc,
                scope_3_ghgp_emissions_source,
                scope_3_ghgp_emissions_reported,
                scope_3_ghgp_emissions_units,
            ) = await self._handle_process_methods(
                process_scope_emissions,
                d=res_values,
                units=res_units,
                source=source,
                last_updated=last_updated,
                emissions=scope_3_fields,
                session=self.session,
                download_option=self.download_option,
                static_cache=self.static_cache,
                restated=restated_fields,
            )
            scope_3_categories_data = {}
            scope_3_categories_desc = {}
            scope_3_categories_source = {}
            scope_3_categories_reported = {}
            scope_3_categories_units = {}
            for i in range(1, 17):
                format_category = "other" if i == 16 else f"c{i}"
                # process scope 3 ghgp category emissions
                scope_3_category_fields = scope_emissions_formatter(
                    res_values,
                    (
                        "s3_emissions_others"
                        if i == 16
                        else "s3_emissions_category"
                    ),
                    "ghgp",
                    format_category,
                )
                (
                    scope_3_ghgp_category_data,
                    scope_3_ghgp_category_desc,
                    scope_3_ghgp_category_source,
                    scope_3_ghgp_category_reported,
                    scope_3_ghgp_category_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process scope 3 ghgp category methodology
                (
                    scope_3_ghgp_methodology_data,
                    scope_3_ghgp_methodology_desc,
                    scope_3_ghgp_methodology_source,
                    scope_3_ghgp_methodology_reported,
                    scope_3_ghgp_methodology_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        f"s3_ghgp_{format_category}_emissions_method_dict"
                    ),
                    units_list=res_units.get(
                        f"s3_ghgp_{format_category}_emissions_method_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name=f"s3_ghgp_{format_category}_emissions_method_dict",
                    session=self.session,
                    static_cache=self.static_cache,
                    download_option=self.download_option,
                    restated=restated_fields,
                )
                # process scope 3 ghgp category emissions disclose
                scope_3_category_exc_fields = scope_emissions_formatter(
                    res_values,
                    "s3_emissions_category_exclusion",
                    "ghgp",
                    format_category,
                )
                (
                    scope_3_ghgp_category_disclose_data,
                    scope_3_ghgp_category_disclose_desc,
                    scope_3_ghgp_category_disclose_source,
                    scope_3_ghgp_category_disclose_reported,
                    scope_3_ghgp_category_disclose_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_exc_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process scope 3 ghgp category exclusion
                (
                    scope_3_ghgp_exclusion_data,
                    scope_3_ghgp_exclusion_desc,
                    scope_3_ghgp_exclusion_source,
                    scope_3_ghgp_exclusion_reported,
                    scope_3_ghgp_exclusion_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        f"s3_ghgp_{format_category}_emissions_exclusion_dict"
                    ),
                    units_list=res_units.get(
                        f"s3_ghgp_{format_category}_emissions_exclusion_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name=f"s3_ghgp_{format_category}_emissions_exclusion_dict",
                    session=self.session,
                    static_cache=self.static_cache,
                    download_option=self.download_option,
                    restated=restated_fields,
                )
                # process scope 3 ghgp category emissions change type
                scope_3_category_change_fields = scope_emissions_formatter(
                    res_values,
                    "s3_emissions_category_change_type",
                    "ghgp",
                    format_category,
                )
                (
                    scope_3_ghgp_category_change_data,
                    scope_3_ghgp_category_change_desc,
                    scope_3_ghgp_category_change_source,
                    scope_3_ghgp_category_change_reported,
                    scope_3_ghgp_category_change_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_change_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # data
                scope_3_categories_data[
                    f"scope_3_ghgp_{format_category}_full"
                ] = scope_3_ghgp_category_data
                scope_3_categories_data[
                    f"scope_3_ghgp_{format_category}_method_full"
                ] = scope_3_ghgp_methodology_data
                scope_3_categories_data[
                    f"scope_3_ghgp_{format_category}_disclose_full"
                ] = scope_3_ghgp_category_disclose_data
                scope_3_categories_data[
                    f"scope_3_ghgp_{format_category}_exclusion_full"
                ] = scope_3_ghgp_exclusion_data
                scope_3_categories_data[
                    f"scope_3_ghgp_{format_category}_change_full"
                ] = scope_3_ghgp_category_change_data
                # description
                scope_3_categories_desc[
                    f"scope_3_ghgp_{format_category}_full"
                ] = scope_3_ghgp_category_desc
                scope_3_categories_desc[
                    f"scope_3_ghgp_{format_category}_method_full"
                ] = scope_3_ghgp_methodology_desc
                scope_3_categories_desc[
                    f"scope_3_ghgp_{format_category}_disclose_full"
                ] = scope_3_ghgp_category_disclose_desc
                scope_3_categories_desc[
                    f"scope_3_ghgp_{format_category}_exclusion_full"
                ] = scope_3_ghgp_exclusion_desc
                scope_3_categories_desc[
                    f"scope_3_ghgp_{format_category}_change_full"
                ] = scope_3_ghgp_category_change_desc
                # source
                scope_3_categories_source[
                    f"scope_3_ghgp_{format_category}_full"
                ] = scope_3_ghgp_category_source
                scope_3_categories_source[
                    f"scope_3_ghgp_{format_category}_method_full"
                ] = scope_3_ghgp_methodology_source
                scope_3_categories_source[
                    f"scope_3_ghgp_{format_category}_disclose_full"
                ] = scope_3_ghgp_category_disclose_source
                scope_3_categories_source[
                    f"scope_3_ghgp_{format_category}_exclusion_full"
                ] = scope_3_ghgp_exclusion_source
                scope_3_categories_source[
                    f"scope_3_ghgp_{format_category}_change_full"
                ] = scope_3_ghgp_category_change_source
                # reported
                scope_3_categories_reported[
                    f"scope_3_ghgp_{format_category}_full"
                ] = scope_3_ghgp_category_reported
                scope_3_categories_reported[
                    f"scope_3_ghgp_{format_category}_method_full"
                ] = scope_3_ghgp_methodology_reported
                scope_3_categories_reported[
                    f"scope_3_ghgp_{format_category}_disclose_full"
                ] = scope_3_ghgp_category_disclose_reported
                scope_3_categories_reported[
                    f"scope_3_ghgp_{format_category}_exclusion_full"
                ] = scope_3_ghgp_exclusion_reported
                scope_3_categories_reported[
                    f"scope_3_ghgp_{format_category}_change_full"
                ] = scope_3_ghgp_category_change_reported
                # units
                scope_3_categories_units[
                    f"scope_3_ghgp_{format_category}_full"
                ] = scope_3_ghgp_category_units
                scope_3_categories_units[
                    f"scope_3_ghgp_{format_category}_method_full"
                ] = scope_3_ghgp_methodology_units
                scope_3_categories_units[
                    f"scope_3_ghgp_{format_category}_disclose_full"
                ] = scope_3_ghgp_category_disclose_units
                scope_3_categories_units[
                    f"scope_3_ghgp_{format_category}_exclusion_full"
                ] = scope_3_ghgp_exclusion_units
                scope_3_categories_units[
                    f"scope_3_ghgp_{format_category}_change_full"
                ] = scope_3_ghgp_category_change_units
            # process scope 3 iso emissions
            scope_3_iso_categories_data = {}
            scope_3_iso_categories_desc = {}
            scope_3_iso_categories_source = {}
            scope_3_iso_categories_reported = {}
            scope_3_iso_categories_units = {}
            for i in range(3, 7):
                format_category = f"c{i}"
                # process scope 3 iso category emissions
                scope_3_category_fields = scope_emissions_formatter(
                    res_values,
                    "s3_emissions_category",
                    "iso",
                    format_category,
                )
                (
                    scope_3_iso_category_data,
                    scope_3_iso_category_desc,
                    scope_3_iso_category_source,
                    scope_3_iso_category_reported,
                    scope_3_iso_category_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process scope 3 iso category methodology
                (
                    scope_3_iso_methodology_data,
                    scope_3_iso_methodology_desc,
                    scope_3_iso_methodology_source,
                    scope_3_iso_methodology_reported,
                    scope_3_iso_methodology_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        f"s3_iso_{format_category}_emissions_method_dict"
                    ),
                    units_list=res_units.get(
                        f"s3_iso_{format_category}_emissions_method_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name=f"s3_iso_{format_category}_emissions_method_dict",
                    session=self.session,
                    static_cache=self.static_cache,
                    download_option=self.download_option,
                    restated=restated_fields,
                )
                # process scope 3 iso category emissions disclose
                scope_3_category_exc_fields = scope_emissions_formatter(
                    res_values,
                    "s3_emissions_category_exclusion",
                    "iso",
                    format_category,
                )
                (
                    scope_3_iso_category_disclose_data,
                    scope_3_iso_category_disclose_desc,
                    scope_3_iso_category_disclose_source,
                    scope_3_iso_category_disclose_reported,
                    scope_3_iso_category_disclose_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_exc_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process scope 3 iso category exclusion
                (
                    scope_3_iso_exclusion_data,
                    scope_3_iso_exclusion_desc,
                    scope_3_iso_exclusion_source,
                    scope_3_iso_exclusion_reported,
                    scope_3_iso_exclusion_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        f"s3_iso_{format_category}_emissions_exclusion_dict"
                    ),
                    units_list=res_units.get(
                        f"s3_iso_{format_category}_emissions_exclusion_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name=f"s3_iso_{format_category}_emissions_exclusion_dict",
                    session=self.session,
                    static_cache=self.static_cache,
                    download_option=self.download_option,
                    restated=restated_fields,
                )
                # process scope 3 iso category emissions change type
                scope_3_category_change_fields = scope_emissions_formatter(
                    res_values,
                    "s3_emissions_category_change_type",
                    "iso",
                    format_category,
                )
                (
                    scope_3_iso_category_change_data,
                    scope_3_iso_category_change_desc,
                    scope_3_iso_category_change_source,
                    scope_3_iso_category_change_reported,
                    scope_3_iso_category_change_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=scope_3_category_change_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # data
                scope_3_iso_categories_data[
                    f"scope_3_iso_{format_category}_full"
                ] = scope_3_iso_category_data
                scope_3_iso_categories_data[
                    f"scope_3_iso_{format_category}_method_full"
                ] = scope_3_iso_methodology_data
                scope_3_iso_categories_data[
                    f"scope_3_iso_{format_category}_disclose_full"
                ] = scope_3_iso_category_disclose_data
                scope_3_iso_categories_data[
                    f"scope_3_iso_{format_category}_exclusion_full"
                ] = scope_3_iso_exclusion_data
                scope_3_iso_categories_data[
                    f"scope_3_iso_{format_category}_change_full"
                ] = scope_3_iso_category_change_data
                # description
                scope_3_iso_categories_desc[
                    f"scope_3_iso_{format_category}_full"
                ] = scope_3_iso_category_desc
                scope_3_iso_categories_desc[
                    f"scope_3_iso_{format_category}_method_full"
                ] = scope_3_iso_methodology_desc
                scope_3_iso_categories_desc[
                    f"scope_3_iso_{format_category}_disclose_full"
                ] = scope_3_iso_category_disclose_desc
                scope_3_iso_categories_desc[
                    f"scope_3_iso_{format_category}_exclusion_full"
                ] = scope_3_iso_exclusion_desc
                scope_3_iso_categories_desc[
                    f"scope_3_iso_{format_category}_change_full"
                ] = scope_3_iso_category_change_desc
                # source
                scope_3_iso_categories_source[
                    f"scope_3_iso_{format_category}_full"
                ] = scope_3_iso_category_source
                scope_3_iso_categories_source[
                    f"scope_3_iso_{format_category}_method_full"
                ] = scope_3_iso_methodology_source
                scope_3_iso_categories_source[
                    f"scope_3_iso_{format_category}_disclose_full"
                ] = scope_3_iso_category_disclose_source
                scope_3_iso_categories_source[
                    f"scope_3_iso_{format_category}_exclusion_full"
                ] = scope_3_iso_exclusion_source
                scope_3_iso_categories_source[
                    f"scope_3_iso_{format_category}_change_full"
                ] = scope_3_iso_category_change_source
                # reported
                scope_3_iso_categories_reported[
                    f"scope_3_iso_{format_category}_full"
                ] = scope_3_iso_category_reported
                scope_3_iso_categories_reported[
                    f"scope_3_iso_{format_category}_method_full"
                ] = scope_3_iso_methodology_reported
                scope_3_iso_categories_reported[
                    f"scope_3_iso_{format_category}_disclose_full"
                ] = scope_3_iso_category_disclose_reported
                scope_3_iso_categories_reported[
                    f"scope_3_iso_{format_category}_exclusion_full"
                ] = scope_3_iso_exclusion_reported
                scope_3_iso_categories_reported[
                    f"scope_3_iso_{format_category}_change_full"
                ] = scope_3_iso_category_change_reported
                # units
                scope_3_iso_categories_units[
                    f"scope_3_iso_{format_category}_full"
                ] = scope_3_iso_category_units
                scope_3_iso_categories_units[
                    f"scope_3_iso_{format_category}_method_full"
                ] = scope_3_iso_methodology_units
                scope_3_iso_categories_units[
                    f"scope_3_iso_{format_category}_disclose_full"
                ] = scope_3_iso_category_disclose_units
                scope_3_iso_categories_units[
                    f"scope_3_iso_{format_category}_exclusion_full"
                ] = scope_3_iso_exclusion_units
                scope_3_iso_categories_units[
                    f"scope_3_iso_{format_category}_change_full"
                ] = scope_3_iso_category_change_units
            # process assure verif
            assure_verif_data = await self._handle_process_methods(
                process_assure_verif_companies,
                d=res_values,
                default_headers=assure_verif_headers,
                reporting_year=reporting_year,
                last_updated=last_updated,
                source_set=source_set,
                session=self.session,
                restated=restated_fields,
            )
            # append assurance and verification total data
            assure_verif_temp.update({reporting_year: assure_verif_data})
            # check for company type if Financial then process it
            if self.company.company_type == "Financial":
                # process financed emissions aum root data
                fe_aum_overview_fields = financed_emissions_formatter(
                    res_values, "fe_emissions", "aum"
                )
                (
                    fe_overview_data,
                    fe_overview_desc,
                    fe_overview_source,
                    fe_overview_reported,
                    fe_overview_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=fe_aum_overview_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions coverage aum data
                (
                    fe_aum_absolute_data,
                    fe_aum_absolute_desc,
                    fe_aum_absolute_source,
                    fe_aum_absolute_reported,
                    fe_aum_absolute_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get("fin_emissions_aum_dict"),
                    units_list=res_units.get("fin_emissions_aum_dict"),
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_aum_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions aum change types data
                fe_aum_hange_type_fields = financed_emissions_formatter(
                    res_values, "fe_change_type", "aum"
                )
                (
                    fe_change_data,
                    fe_change_desc,
                    fe_change_source,
                    fe_change_reported,
                    fe_change_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=fe_aum_hange_type_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions intensity aum data
                # remove _units and _units_usd fields from fe
                new_fe = remove_fields_from_form(
                    values=res_values.get("fin_emissions_int_aum_dict"),
                    fields=[
                        "fin_emissions_int_aum_units",
                        "fin_emissions_int_aum_units_usd",
                    ],
                )
                new_units_fe = (
                    remove_fields_from_form(
                        values=res_units.get("fin_emissions_int_aum_dict"),
                        fields=[
                            "fin_emissions_int_aum_units",
                            "fin_emissions_int_aum_units_usd",
                        ],
                    )
                    if res_units.get("fin_emissions_int_aum_dict") is not None
                    else None
                )
                (
                    fe_aum_intensity_data,
                    fe_aum_intensity_desc,
                    fe_aum_intensity_source,
                    fe_aum_intensity_reported,
                    fe_aum_intensity_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=new_fe,
                    units_list=new_units_fe,
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_int_aum_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions data quality aum data
                (
                    fe_aum_data_quality_data,
                    fe_aum_data_quality_desc,
                    fe_aum_data_quality_source,
                    fe_aum_data_quality_reported,
                    fe_aum_data_quality_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        "fin_emissions_aum_data_quality_dict"
                    ),
                    units_list=res_units.get(
                        "fin_emissions_aum_data_quality_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_aum_data_quality_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions grossexp root data
                fe_ge_overview_fields = financed_emissions_formatter(
                    res_values, "fe_emissions", "grossexp"
                )
                (
                    fe_ge_overview_data,
                    fe_ge_overview_desc,
                    fe_ge_overview_source,
                    fe_ge_overview_reported,
                    fe_ge_overview_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=fe_ge_overview_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions coverage grossexp data
                (
                    fe_ge_absolute_data,
                    fe_ge_absolute_desc,
                    fe_ge_absolute_source,
                    fe_ge_absolute_reported,
                    fe_ge_absolute_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        "fin_emissions_grossexp_dict"
                    ),
                    units_list=res_units.get("fin_emissions_grossexp_dict"),
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_grossexp_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions grossexp change types data
                fe_ge_hange_type_fields = financed_emissions_formatter(
                    res_values, "fe_change_type", "grossexp"
                )
                (
                    fe_ge_change_data,
                    fe_ge_change_desc,
                    fe_ge_change_source,
                    fe_ge_change_reported,
                    fe_ge_change_units,
                ) = await self._handle_process_methods(
                    process_scope_emissions,
                    d=res_values,
                    units=res_units,
                    source=source,
                    last_updated=last_updated,
                    emissions=fe_ge_hange_type_fields,
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions intensity grossexp data
                # remove _units and _units_usd fields from fe

                new_ge_fe = remove_fields_from_form(
                    values=res_values.get("fin_emissions_int_grossexp_dict"),
                    fields=[
                        "fin_emissions_int_grossexp_units",
                        "fin_emissions_int_grossexp_units_usd",
                    ],
                )
                new_units_ge_fe = (
                    remove_fields_from_form(
                        values=res_units.get(
                            "fin_emissions_int_grossexp_dict"
                        ),
                        fields=[
                            "fin_emissions_int_grossexp_units",
                            "fin_emissions_int_grossexp_units_usd",
                        ],
                    )
                    if res_units.get("fin_emissions_int_grossexp_dict")
                    is not None
                    else None
                )
                (
                    fe_ge_intensity_data,
                    fe_ge_intensity_desc,
                    fe_ge_intensity_source,
                    fe_ge_intensity_reported,
                    fe_ge_intensity_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=new_ge_fe,
                    units_list=new_units_ge_fe,
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_int_grossexp_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
                # process financed emissions data quality grossexp data
                (
                    fe_ge_data_quality_data,
                    fe_ge_data_quality_desc,
                    fe_ge_data_quality_source,
                    fe_ge_data_quality_reported,
                    fe_ge_data_quality_units,
                ) = await self._handle_process_methods(
                    process_data_export,
                    data_export_list=res_values.get(
                        "fin_emissions_grossexp_data_quality_dict"
                    ),
                    units_list=res_units.get(
                        "fin_emissions_grossexp_data_quality_dict"
                    ),
                    source=source,
                    last_updated=last_updated,
                    form_name="fin_emissions_grossexp_data_quality_dict",
                    session=self.session,
                    download_option=self.download_option,
                    static_cache=self.static_cache,
                    restated=restated_fields,
                )
            target_ids = {}
            targets_abs_data = []
            targets_int_data = []
            targets_abs_data_progress = []
            targets_int_data_progress = []
            # process absolute targets data
            tgt_abs_sub_form = res_values.get("tgt_abs_dict")
            if isinstance(tgt_abs_sub_form, list) and tgt_abs_sub_form[-1].get(
                "tgt_abs_id"
            ) not in (NullTypeState.LONG_DASH.value, None):
                (
                    targets_abs_data,
                    tgt_abs_ids,
                ) = await self._handle_process_methods(
                    process_targets_export,
                    data_export_list=res_values.get("tgt_abs_dict"),
                    source=source,
                    last_updated=last_updated,
                    units_list=res_units.get("tgt_abs_dict"),
                    form_name="tgt_abs_dict",
                    session=self.session,
                    reporting_year=reporting_year,
                    disclosure_year=disclosure_year,
                    target_ids=target_ids,
                )
                target_ids.update(tgt_abs_ids)
            # Updated Excel generation logic with additional check for tgt_int_intensity_type
            tgt_int_sub_form = res_values.get("tgt_int_dict")
            if isinstance(tgt_int_sub_form, list) and tgt_int_sub_form[-1].get(
                "tgt_int_id"
            ) not in (NullTypeState.LONG_DASH.value, None):
                # Check for intensity type
                intensity_type = (
                    tgt_int_sub_form[-1]
                    .get("tgt_int_intensity_type", "")
                    .strip()
                    .lower()
                )
                is_physical_intensity = "physical" in intensity_type

                if is_physical_intensity:
                    # List of columns to exclude for physical intensity
                    exclude_columns = ["metric_usd", "int_usd"]
                    # Filter the tgt_int_sub_form to exclude unwanted columns
                    filtered_tgt_int_sub_form = [
                        {
                            key: scientific_to_float(value)
                            for key, value in row.items()
                            if not any(col in key for col in exclude_columns)
                        }
                        for row in tgt_int_sub_form
                    ]
                else:
                    # No exclusion for economic intensity
                    filtered_tgt_int_sub_form = tgt_int_sub_form

                (
                    targets_int_data,
                    tgt_int_ids,
                ) = await self._handle_process_methods(
                    process_targets_export,
                    data_export_list=filtered_tgt_int_sub_form,
                    source=source,
                    last_updated=last_updated,
                    units_list=res_units.get("tgt_int_dict"),
                    form_name="tgt_int_dict",
                    session=self.session,
                    reporting_year=reporting_year,
                    disclosure_year=disclosure_year,
                    target_ids=target_ids,
                )
                target_ids.update(tgt_int_ids)
            target_combine = targets_abs_data + targets_int_data
            targets_full_list[0:0] = target_combine

            # process absolute targets progress data
            tgt_abs_progress_sub_form = res_values.get("tgt_abs_progress_dict")
            if isinstance(
                tgt_abs_progress_sub_form, list
            ) and tgt_abs_progress_sub_form[-1].get(
                "tgt_abs_id_progress"
            ) not in (
                NullTypeState.LONG_DASH.value,
                None,
            ):
                # combine target and target progress for units
                target_units = combine_units_into_one_list(
                    root_list=res_units.get("tgt_abs_dict"),
                    sub_list=res_units.get("tgt_abs_progress_dict"),
                )
                (
                    targets_abs_data_progress,
                    _,
                ) = await self._handle_process_methods(
                    process_targets_export,
                    data_export_list=res_values.get("tgt_abs_progress_dict"),
                    source=source,
                    last_updated=last_updated,
                    units_list=target_units,
                    form_name="tgt_abs_progress_dict",
                    session=self.session,
                    reporting_year=reporting_year,
                    disclosure_year=disclosure_year,
                    target_ids=target_ids,
                    progress=True,
                )

            # Process intensity targets progress data
            tgt_int_progress_sub_form = res_values.get("tgt_int_progress_dict")
            if isinstance(
                tgt_int_progress_sub_form, list
            ) and tgt_int_progress_sub_form[-1].get(
                "tgt_int_id_progress"
            ) not in (
                NullTypeState.LONG_DASH.value,
                None,
            ):
                intensity_type = (
                    (
                        tgt_int_sub_form[-1]
                        .get("tgt_int_intensity_type", "")
                        .strip()
                        .lower()
                    )
                    if isinstance(tgt_int_sub_form, list)
                    else ""
                )
                is_physical_intensity = "physical" in intensity_type

                if is_physical_intensity:
                    # List of columns to exclude
                    exclude_columns = ["metric_usd", "int_usd"]
                    # Filter the tgt_int_progress_sub_form to exclude unwanted columns
                    filtered_tgt_int_progress_sub_form = [
                        {
                            key: scientific_to_float(value)
                            for key, value in row.items()
                            if not any(col in key for col in exclude_columns)
                        }
                        for row in tgt_int_progress_sub_form
                    ]
                else:
                    # No exclusion for economic intensity
                    filtered_tgt_int_progress_sub_form = (
                        tgt_int_progress_sub_form
                    )

                # Combine target and target progress for units
                target_units = combine_units_into_one_list(
                    root_list=res_units.get("tgt_int_dict"),
                    sub_list=res_units.get("tgt_int_progress_dict"),
                )
                (
                    targets_int_data_progress,
                    _,
                ) = await self._handle_process_methods(
                    process_targets_export,
                    data_export_list=filtered_tgt_int_progress_sub_form,
                    source=source,
                    last_updated=last_updated,
                    units_list=target_units,
                    form_name="tgt_int_progress_dict",
                    session=self.session,
                    reporting_year=reporting_year,
                    disclosure_year=disclosure_year,
                    target_ids=target_ids,
                    progress=True,
                )
            target_progress_combine = (
                targets_abs_data_progress + targets_int_data_progress
            )
            targets_progress_full_list[0:0] = target_progress_combine

            # process target validation data
            validation_abs_data = await self._handle_process_methods(
                process_target_validation_companies,
                d=res_values,
                disclosure_year=disclosure_year,
                last_updated=last_updated,
                form_name="tgt_abs_valid_dict",
                session=self.session,
                target_ids=target_ids,
            )
            validation_int_data = await self._handle_process_methods(
                process_target_validation_companies,
                d=res_values,
                disclosure_year=disclosure_year,
                last_updated=last_updated,
                form_name="tgt_int_valid_dict",
                session=self.session,
                target_ids=target_ids,
            )
            validation_target_progress_combine = (
                validation_abs_data + validation_int_data
            )
            # append assurance and verification total data
            validation_full_list[0:0] = validation_target_progress_combine
            # create dataframes for metadata
            metadata_df = {"company_metadata_full": company_metadata_full}
            metadata_data = {"company_metadata_full": company_metadata_data}
            metadata_desc = {"company_metadata_full": company_metadata_desc}
            # create dataframes for scope emissions
            dataframes = {
                key: dataframes_scope_results[key] for key in scope_results
            }
            # create data all together
            data = {
                "scope_emissions_full": scope_emissions_data,
                "scope_exclusion_full": scope_exclusion_data,
                "scope_emissions_change_full": scope_change_type_data,
                "scope_ghg_full": scope_ghg_data,
                "scope_ghg_other_full": scope_ghg_other_data,
                "scope_lb_emissions_full": scope_lb_emissions_data,
                "scope_lb_exclusion_full": scope_lb_exclusion_data,
                "scope_lb_change_full": scope_lb_change_type_data,
                "scope_mb_emissions_full": scope_mb_emissions_data,
                "scope_mb_exclusion_full": scope_mb_exclusion_data,
                "scope_mb_change_full": scope_mb_change_type_data,
                "scope_3_ghgp_emissions_full": scope_3_ghgp_emissions_data,
            }
            # iterate over scope 3 to format dataframes data
            for i in range(1, 16):
                data[f"scope_3_ghgp_c{i}_full"] = scope_3_categories_data.get(
                    f"scope_3_ghgp_c{i}_full", {}
                )
                data[f"scope_3_ghgp_c{i}_method_full"] = (
                    scope_3_categories_data.get(
                        f"scope_3_ghgp_c{i}_method_full", {}
                    )
                )
                data[f"scope_3_ghgp_c{i}_disclose_full"] = (
                    scope_3_categories_data.get(
                        f"scope_3_ghgp_c{i}_disclose_full", {}
                    )
                )
                data[f"scope_3_ghgp_c{i}_exclusion_full"] = (
                    scope_3_categories_data.get(
                        f"scope_3_ghgp_c{i}_exclusion_full", {}
                    )
                )
                data[f"scope_3_ghgp_c{i}_change_full"] = (
                    scope_3_categories_data.get(
                        f"scope_3_ghgp_c{i}_change_full", {}
                    )
                )

            data["scope_3_ghgp_other_full"] = scope_3_categories_data.get(
                "scope_3_ghgp_other_full"
            )
            data["scope_3_ghgp_other_method_full"] = (
                scope_3_categories_data.get("scope_3_ghgp_other_method_full")
            )
            data["scope_3_ghgp_other_disclose_full"] = (
                scope_3_categories_data.get("scope_3_ghgp_other_disclose_full")
            )
            data["scope_3_ghgp_other_exclusion_full"] = (
                scope_3_categories_data.get(
                    "scope_3_ghgp_other_exclusion_full"
                )
            )
            data["scope_3_ghgp_other_change_full"] = (
                scope_3_categories_data.get("scope_3_ghgp_other_change_full")
            )
            for i in range(3, 7):
                data[f"scope_3_iso_c{i}_full"] = (
                    scope_3_iso_categories_data.get(
                        f"scope_3_iso_c{i}_full", {}
                    )
                )
                data[f"scope_3_iso_c{i}_method_full"] = (
                    scope_3_iso_categories_data.get(
                        f"scope_3_iso_c{i}_method_full", {}
                    )
                )
                data[f"scope_3_iso_c{i}_disclose_full"] = (
                    scope_3_iso_categories_data.get(
                        f"scope_3_iso_c{i}_disclose_full", {}
                    )
                )
                data[f"scope_3_iso_c{i}_exclusion_full"] = (
                    scope_3_iso_categories_data.get(
                        f"scope_3_iso_c{i}_exclusion_full", {}
                    )
                )
                data[f"scope_3_iso_c{i}_change_full"] = (
                    scope_3_iso_categories_data.get(
                        f"scope_3_iso_c{i}_change_full", {}
                    )
                )
            # create description all together
            desc = {
                "scope_emissions_full": scope_emissions_desc,
                "scope_exclusion_full": scope_exclusion_desc,
                "scope_emissions_change_full": scope_change_type_desc,
                "scope_ghg_full": scope_ghg_desc,
                "scope_ghg_other_full": scope_ghg_other_desc,
                "scope_lb_emissions_full": scope_lb_emissions_desc,
                "scope_lb_exclusion_full": scope_lb_exclusion_desc,
                "scope_lb_change_full": scope_lb_change_type_desc,
                "scope_mb_emissions_full": scope_mb_emissions_desc,
                "scope_mb_exclusion_full": scope_mb_exclusion_desc,
                "scope_mb_change_full": scope_mb_change_type_desc,
                "scope_3_ghgp_emissions_full": scope_3_ghgp_emissions_desc,
            }
            # iterate over scope 3 to format dataframes data
            for i in range(1, 16):
                desc[f"scope_3_ghgp_c{i}_full"] = scope_3_categories_desc.get(
                    f"scope_3_ghgp_c{i}_full", {}
                )
                desc[f"scope_3_ghgp_c{i}_method_full"] = (
                    scope_3_categories_desc.get(
                        f"scope_3_ghgp_c{i}_method_full", {}
                    )
                )
                desc[f"scope_3_ghgp_c{i}_disclose_full"] = (
                    scope_3_categories_desc.get(
                        f"scope_3_ghgp_c{i}_disclose_full", {}
                    )
                )
                desc[f"scope_3_ghgp_c{i}_exclusion_full"] = (
                    scope_3_categories_desc.get(
                        f"scope_3_ghgp_c{i}_exclusion_full", {}
                    )
                )
                desc[f"scope_3_ghgp_c{i}_change_full"] = (
                    scope_3_categories_desc.get(
                        f"scope_3_ghgp_c{i}_change_full", {}
                    )
                )

            desc["scope_3_ghgp_other_full"] = scope_3_categories_desc.get(
                "scope_3_ghgp_other_full"
            )
            desc["scope_3_ghgp_other_method_full"] = (
                scope_3_categories_desc.get("scope_3_ghgp_other_method_full")
            )
            desc["scope_3_ghgp_other_disclose_full"] = (
                scope_3_categories_desc.get("scope_3_ghgp_other_disclose_full")
            )
            desc["scope_3_ghgp_other_exclusion_full"] = (
                scope_3_categories_desc.get(
                    "scope_3_ghgp_other_exclusion_full"
                )
            )
            desc["scope_3_ghgp_other_change_full"] = (
                scope_3_categories_desc.get("scope_3_ghgp_other_change_full")
            )
            for i in range(3, 7):
                desc[f"scope_3_iso_c{i}_full"] = (
                    scope_3_iso_categories_desc.get(
                        f"scope_3_iso_c{i}_full", {}
                    )
                )
                desc[f"scope_3_iso_c{i}_method_full"] = (
                    scope_3_iso_categories_desc.get(
                        f"scope_3_iso_c{i}_method_full", {}
                    )
                )
                desc[f"scope_3_iso_c{i}_disclose_full"] = (
                    scope_3_iso_categories_desc.get(
                        f"scope_3_iso_c{i}_disclose_full", {}
                    )
                )
                desc[f"scope_3_iso_c{i}_exclusion_full"] = (
                    scope_3_iso_categories_desc.get(
                        f"scope_3_iso_c{i}_exclusion_full", {}
                    )
                )
                desc[f"scope_3_iso_c{i}_change_full"] = (
                    scope_3_iso_categories_desc.get(
                        f"scope_3_iso_c{i}_change_full", {}
                    )
                )
            # create source all together
            source_data = {
                "scope_emissions_full": scope_emissions_source,
                "scope_exclusion_full": scope_exclusion_source,
                "scope_emissions_change_full": scope_change_type_source,
                "scope_ghg_full": scope_ghg_source,
                "scope_ghg_other_full": scope_ghg_other_source,
                "scope_lb_emissions_full": scope_lb_emissions_source,
                "scope_lb_exclusion_full": scope_lb_exclusion_source,
                "scope_lb_change_full": scope_lb_change_type_source,
                "scope_mb_emissions_full": scope_mb_emissions_source,
                "scope_mb_exclusion_full": scope_mb_exclusion_source,
                "scope_mb_change_full": scope_mb_change_type_source,
                "scope_3_ghgp_emissions_full": scope_3_ghgp_emissions_source,
            }
            # iterate over scope 3 to format dataframes data
            for i in range(1, 16):
                source_data[f"scope_3_ghgp_c{i}_full"] = (
                    scope_3_categories_source.get(
                        f"scope_3_ghgp_c{i}_full", {}
                    )
                )
                source_data[f"scope_3_ghgp_c{i}_method_full"] = (
                    scope_3_categories_source.get(
                        f"scope_3_ghgp_c{i}_method_full", {}
                    )
                )
                source_data[f"scope_3_ghgp_c{i}_disclose_full"] = (
                    scope_3_categories_source.get(
                        f"scope_3_ghgp_c{i}_disclose_full", {}
                    )
                )
                source_data[f"scope_3_ghgp_c{i}_exclusion_full"] = (
                    scope_3_categories_source.get(
                        f"scope_3_ghgp_c{i}_exclusion_full", {}
                    )
                )
                source_data[f"scope_3_ghgp_c{i}_change_full"] = (
                    scope_3_categories_source.get(
                        f"scope_3_ghgp_c{i}_change_full", {}
                    )
                )

            source_data["scope_3_ghgp_other_full"] = (
                scope_3_categories_source.get("scope_3_ghgp_other_full")
            )
            source_data["scope_3_ghgp_other_method_full"] = (
                scope_3_categories_source.get("scope_3_ghgp_other_method_full")
            )
            source_data["scope_3_ghgp_other_disclose_full"] = (
                scope_3_categories_source.get(
                    "scope_3_ghgp_other_disclose_full"
                )
            )
            source_data["scope_3_ghgp_other_exclusion_full"] = (
                scope_3_categories_source.get(
                    "scope_3_ghgp_other_exclusion_full"
                )
            )
            source_data["scope_3_ghgp_other_change_full"] = (
                scope_3_categories_source.get("scope_3_ghgp_other_change_full")
            )
            for i in range(3, 7):
                source_data[f"scope_3_iso_c{i}_full"] = (
                    scope_3_iso_categories_source.get(
                        f"scope_3_iso_c{i}_full", {}
                    )
                )
                source_data[f"scope_3_iso_c{i}_method_full"] = (
                    scope_3_iso_categories_source.get(
                        f"scope_3_iso_c{i}_method_full", {}
                    )
                )
                source_data[f"scope_3_iso_c{i}_disclose_full"] = (
                    scope_3_iso_categories_source.get(
                        f"scope_3_iso_c{i}_disclose_full", {}
                    )
                )
                source_data[f"scope_3_iso_c{i}_exclusion_full"] = (
                    scope_3_iso_categories_source.get(
                        f"scope_3_iso_c{i}_exclusion_full", {}
                    )
                )
                source_data[f"scope_3_iso_c{i}_change_full"] = (
                    scope_3_iso_categories_source.get(
                        f"scope_3_iso_c{i}_change_full", {}
                    )
                )
            # create last_reported all together
            reported_data = {
                "scope_emissions_full": scope_emissions_reported,
                "scope_exclusion_full": scope_exclusion_reported,
                "scope_emissions_change_full": scope_change_type_reported,
                "scope_ghg_full": scope_ghg_reported,
                "scope_ghg_other_full": scope_ghg_other_reported,
                "scope_lb_emissions_full": scope_lb_emissions_reported,
                "scope_lb_exclusion_full": scope_lb_exclusion_reported,
                "scope_lb_change_full": scope_lb_change_type_reported,
                "scope_mb_emissions_full": scope_mb_emissions_reported,
                "scope_mb_exclusion_full": scope_mb_exclusion_reported,
                "scope_mb_change_full": scope_mb_change_type_reported,
                "scope_3_ghgp_emissions_full": scope_3_ghgp_emissions_reported,
            }
            # iterate over scope 3 to format dataframes data
            for i in range(1, 16):
                reported_data[f"scope_3_ghgp_c{i}_full"] = (
                    scope_3_categories_reported.get(
                        f"scope_3_ghgp_c{i}_full", {}
                    )
                )
                reported_data[f"scope_3_ghgp_c{i}_method_full"] = (
                    scope_3_categories_reported.get(
                        f"scope_3_ghgp_c{i}_method_full", {}
                    )
                )
                reported_data[f"scope_3_ghgp_c{i}_disclose_full"] = (
                    scope_3_categories_reported.get(
                        f"scope_3_ghgp_c{i}_disclose_full", {}
                    )
                )
                reported_data[f"scope_3_ghgp_c{i}_exclusion_full"] = (
                    scope_3_categories_reported.get(
                        f"scope_3_ghgp_c{i}_exclusion_full", {}
                    )
                )
                reported_data[f"scope_3_ghgp_c{i}_change_full"] = (
                    scope_3_categories_reported.get(
                        f"scope_3_ghgp_c{i}_change_full", {}
                    )
                )

            reported_data["scope_3_ghgp_other_full"] = (
                scope_3_categories_reported.get("scope_3_ghgp_other_full")
            )
            reported_data["scope_3_ghgp_other_method_full"] = (
                scope_3_categories_reported.get(
                    "scope_3_ghgp_other_method_full"
                )
            )
            reported_data["scope_3_ghgp_other_disclose_full"] = (
                scope_3_categories_reported.get(
                    "scope_3_ghgp_other_disclose_full"
                )
            )
            reported_data["scope_3_ghgp_other_exclusion_full"] = (
                scope_3_categories_reported.get(
                    "scope_3_ghgp_other_exclusion_full"
                )
            )
            reported_data["scope_3_ghgp_other_change_full"] = (
                scope_3_categories_reported.get(
                    "scope_3_ghgp_other_change_full"
                )
            )
            for i in range(3, 7):
                reported_data[f"scope_3_iso_c{i}_full"] = (
                    scope_3_iso_categories_reported.get(
                        f"scope_3_iso_c{i}_full", {}
                    )
                )
                reported_data[f"scope_3_iso_c{i}_method_full"] = (
                    scope_3_iso_categories_reported.get(
                        f"scope_3_iso_c{i}_method_full", {}
                    )
                )
                reported_data[f"scope_3_iso_c{i}_disclose_full"] = (
                    scope_3_iso_categories_reported.get(
                        f"scope_3_iso_c{i}_disclose_full", {}
                    )
                )
                reported_data[f"scope_3_iso_c{i}_exclusion_full"] = (
                    scope_3_iso_categories_reported.get(
                        f"scope_3_iso_c{i}_exclusion_full", {}
                    )
                )
                reported_data[f"scope_3_iso_c{i}_change_full"] = (
                    scope_3_iso_categories_reported.get(
                        f"scope_3_iso_c{i}_change_full", {}
                    )
                )
            # create units all together
            units_data = {
                "scope_emissions_full": scope_emissions_units,
                "scope_exclusion_full": scope_exclusion_units,
                "scope_emissions_change_full": scope_change_type_units,
                "scope_ghg_full": scope_ghg_units,
                "scope_ghg_other_full": scope_ghg_other_units,
                "scope_lb_emissions_full": scope_lb_emissions_units,
                "scope_lb_exclusion_full": scope_lb_exclusion_units,
                "scope_lb_change_full": scope_lb_change_type_units,
                "scope_mb_emissions_full": scope_mb_emissions_units,
                "scope_mb_exclusion_full": scope_mb_exclusion_units,
                "scope_mb_change_full": scope_mb_change_type_units,
                "scope_3_ghgp_emissions_full": scope_3_ghgp_emissions_units,
            }
            # iterate over scope 3 to format dataframes data
            for i in range(1, 16):
                units_data[f"scope_3_ghgp_c{i}_full"] = (
                    scope_3_categories_units.get(f"scope_3_ghgp_c{i}_full", {})
                )
                units_data[f"scope_3_ghgp_c{i}_method_full"] = (
                    scope_3_categories_units.get(
                        f"scope_3_ghgp_c{i}_method_full", {}
                    )
                )
                units_data[f"scope_3_ghgp_c{i}_disclose_full"] = (
                    scope_3_categories_units.get(
                        f"scope_3_ghgp_c{i}_disclose_full", {}
                    )
                )
                units_data[f"scope_3_ghgp_c{i}_exclusion_full"] = (
                    scope_3_categories_units.get(
                        f"scope_3_ghgp_c{i}_exclusion_full", {}
                    )
                )
                units_data[f"scope_3_ghgp_c{i}_change_full"] = (
                    scope_3_categories_units.get(
                        f"scope_3_ghgp_c{i}_change_full", {}
                    )
                )

            units_data["scope_3_ghgp_other_full"] = (
                scope_3_categories_units.get("scope_3_ghgp_other_full")
            )
            units_data["scope_3_ghgp_other_method_full"] = (
                scope_3_categories_units.get("scope_3_ghgp_other_method_full")
            )
            units_data["scope_3_ghgp_other_disclose_full"] = (
                scope_3_categories_units.get(
                    "scope_3_ghgp_other_disclose_full"
                )
            )
            units_data["scope_3_ghgp_other_exclusion_full"] = (
                scope_3_categories_units.get(
                    "scope_3_ghgp_other_exclusion_full"
                )
            )
            units_data["scope_3_ghgp_other_change_full"] = (
                scope_3_categories_units.get("scope_3_ghgp_other_change_full")
            )
            for i in range(3, 7):
                units_data[f"scope_3_iso_c{i}_full"] = (
                    scope_3_iso_categories_units.get(
                        f"scope_3_iso_c{i}_full", {}
                    )
                )
                units_data[f"scope_3_iso_c{i}_method_full"] = (
                    scope_3_iso_categories_units.get(
                        f"scope_3_iso_c{i}_method_full", {}
                    )
                )
                units_data[f"scope_3_iso_c{i}_disclose_full"] = (
                    scope_3_iso_categories_units.get(
                        f"scope_3_iso_c{i}_disclose_full", {}
                    )
                )
                units_data[f"scope_3_iso_c{i}_exclusion_full"] = (
                    scope_3_iso_categories_units.get(
                        f"scope_3_iso_c{i}_exclusion_full", {}
                    )
                )
                units_data[f"scope_3_iso_c{i}_change_full"] = (
                    scope_3_iso_categories_units.get(
                        f"scope_3_iso_c{i}_change_full", {}
                    )
                )
            # check if company type is Financila to process it
            if self.company.company_type == "Financial":
                fe_df_full = {
                    "fe_aum_overview_full": fe_aum_overview_full,
                    "fe_aum_absolute_full": fe_aum_absolute_full,
                    "fe_aum_change_full": fe_aum_change_full,
                    "fe_aum_intensity_full": fe_aum_intensity_full,
                    "fe_aum_data_quality_full": fe_aum_data_quality_full,
                    "fe_ge_overview_full": fe_ge_overview_full,
                    "fe_ge_absolute_full": fe_ge_absolute_full,
                    "fe_ge_change_full": fe_ge_change_full,
                    "fe_ge_intensity_full": fe_ge_intensity_full,
                    "fe_ge_data_quality_full": fe_ge_data_quality_full,
                }
                fe_df_full_data = {
                    "fe_aum_overview_full": fe_overview_data,
                    "fe_aum_absolute_full": fe_aum_absolute_data,
                    "fe_aum_change_full": fe_change_data,
                    "fe_aum_intensity_full": fe_aum_intensity_data,
                    "fe_aum_data_quality_full": fe_aum_data_quality_data,
                    "fe_ge_overview_full": fe_ge_overview_data,
                    "fe_ge_absolute_full": fe_ge_absolute_data,
                    "fe_ge_change_full": fe_ge_change_data,
                    "fe_ge_intensity_full": fe_ge_intensity_data,
                    "fe_ge_data_quality_full": fe_ge_data_quality_data,
                }
                fe_df_full_desc = {
                    "fe_aum_overview_full": fe_overview_desc,
                    "fe_aum_absolute_full": fe_aum_absolute_desc,
                    "fe_aum_change_full": fe_change_desc,
                    "fe_aum_intensity_full": fe_aum_intensity_desc,
                    "fe_aum_data_quality_full": fe_aum_data_quality_desc,
                    "fe_ge_overview_full": fe_ge_overview_desc,
                    "fe_ge_absolute_full": fe_ge_absolute_desc,
                    "fe_ge_change_full": fe_ge_change_desc,
                    "fe_ge_intensity_full": fe_ge_intensity_desc,
                    "fe_ge_data_quality_full": fe_ge_data_quality_desc,
                }
                fe_df_full_source = {
                    "fe_aum_overview_full": fe_overview_source,
                    "fe_aum_absolute_full": fe_aum_absolute_source,
                    "fe_aum_change_full": fe_change_source,
                    "fe_aum_intensity_full": fe_aum_intensity_source,
                    "fe_aum_data_quality_full": fe_aum_data_quality_source,
                    "fe_ge_overview_full": fe_ge_overview_source,
                    "fe_ge_absolute_full": fe_ge_absolute_source,
                    "fe_ge_change_full": fe_ge_change_source,
                    "fe_ge_intensity_full": fe_ge_intensity_source,
                    "fe_ge_data_quality_full": fe_ge_data_quality_source,
                }
                fe_df_full_reported = {
                    "fe_aum_overview_full": fe_overview_reported,
                    "fe_aum_absolute_full": fe_aum_absolute_reported,
                    "fe_aum_change_full": fe_change_reported,
                    "fe_aum_intensity_full": fe_aum_intensity_reported,
                    "fe_aum_data_quality_full": fe_aum_data_quality_reported,
                    "fe_ge_overview_full": fe_ge_overview_reported,
                    "fe_ge_absolute_full": fe_ge_absolute_reported,
                    "fe_ge_change_full": fe_ge_change_reported,
                    "fe_ge_intensity_full": fe_ge_intensity_reported,
                    "fe_ge_data_quality_full": fe_ge_data_quality_reported,
                }
                fe_df_full_units = {
                    "fe_aum_overview_full": fe_overview_units,
                    "fe_aum_absolute_full": fe_aum_absolute_units,
                    "fe_aum_change_full": fe_change_units,
                    "fe_aum_intensity_full": fe_aum_intensity_units,
                    "fe_aum_data_quality_full": fe_aum_data_quality_units,
                    "fe_ge_overview_full": fe_ge_overview_units,
                    "fe_ge_absolute_full": fe_ge_absolute_units,
                    "fe_ge_change_full": fe_ge_change_units,
                    "fe_ge_intensity_full": fe_ge_intensity_units,
                    "fe_ge_data_quality_full": fe_ge_data_quality_units,
                }
            for k in metadata_df:
                metadata_df[k] = await update_dataframe(
                    metadata_df[k],
                    metadata_data[k],
                    metadata_desc[k],
                    None,
                    None,
                    None,
                    source,
                    last_updated,
                    reporting_year,
                    "metadata",
                    self.session,
                    static_cache=self.static_cache,
                )
            for key in dataframes:
                dataframes[key] = await update_dataframe(
                    dataframes[key],
                    data[key],
                    desc[key],
                    source_data[key],
                    reported_data[key],
                    units_data[key],
                    source,
                    last_updated,
                    reporting_year,
                    "scope",
                    self.session,
                    self.static_cache,
                    restated,
                )
            if self.company.company_type == "Financial":
                for key in fe_df_full:
                    fe_df_full[key] = await update_dataframe(
                        fe_df_full[key],
                        fe_df_full_data[key],
                        fe_df_full_desc[key],
                        fe_df_full_source[key],
                        fe_df_full_reported[key],
                        fe_df_full_units[key],
                        source,
                        last_updated,
                        reporting_year,
                        "fe",
                        self.session,
                        self.static_cache,
                        restated,
                    )
            company_metadata_full = metadata_df["company_metadata_full"]
            scope_emissions_full = dataframes["scope_emissions_full"]
            scope_exclusion_full = dataframes["scope_exclusion_full"]
            scope_emissions_change_full = dataframes[
                "scope_emissions_change_full"
            ]
            scope_ghg_full = dataframes["scope_ghg_full"]
            scope_ghg_other_full = dataframes["scope_ghg_other_full"]
            scope_lb_emissions_full = dataframes["scope_lb_emissions_full"]
            scope_lb_exclusion_full = dataframes["scope_lb_exclusion_full"]
            scope_lb_change_full = dataframes["scope_lb_change_full"]
            scope_mb_emissions_full = dataframes["scope_mb_emissions_full"]
            scope_mb_exclusion_full = dataframes["scope_mb_exclusion_full"]
            scope_mb_change_full = dataframes["scope_mb_change_full"]
            scope_3_ghgp_emissions_full = dataframes[
                "scope_3_ghgp_emissions_full"
            ]
            scope_3_ghgp_c1_full = dataframes["scope_3_ghgp_c1_full"]
            scope_3_ghgp_c1_method_full = dataframes[
                "scope_3_ghgp_c1_method_full"
            ]
            scope_3_ghgp_c1_disclose_full = dataframes[
                "scope_3_ghgp_c1_disclose_full"
            ]
            scope_3_ghgp_c1_exclusion_full = dataframes[
                "scope_3_ghgp_c1_exclusion_full"
            ]
            scope_3_ghgp_c1_change_full = dataframes[
                "scope_3_ghgp_c1_change_full"
            ]
            scope_3_ghgp_c2_full = dataframes["scope_3_ghgp_c2_full"]
            scope_3_ghgp_c2_method_full = dataframes[
                "scope_3_ghgp_c2_method_full"
            ]
            scope_3_ghgp_c2_disclose_full = dataframes[
                "scope_3_ghgp_c2_disclose_full"
            ]
            scope_3_ghgp_c2_exclusion_full = dataframes[
                "scope_3_ghgp_c2_exclusion_full"
            ]
            scope_3_ghgp_c2_change_full = dataframes[
                "scope_3_ghgp_c2_change_full"
            ]
            scope_3_ghgp_c3_full = dataframes["scope_3_ghgp_c3_full"]
            scope_3_ghgp_c3_method_full = dataframes[
                "scope_3_ghgp_c3_method_full"
            ]
            scope_3_ghgp_c3_disclose_full = dataframes[
                "scope_3_ghgp_c3_disclose_full"
            ]
            scope_3_ghgp_c3_exclusion_full = dataframes[
                "scope_3_ghgp_c3_exclusion_full"
            ]
            scope_3_ghgp_c3_change_full = dataframes[
                "scope_3_ghgp_c3_change_full"
            ]
            scope_3_ghgp_c4_full = dataframes["scope_3_ghgp_c4_full"]
            scope_3_ghgp_c4_method_full = dataframes[
                "scope_3_ghgp_c4_method_full"
            ]
            scope_3_ghgp_c4_disclose_full = dataframes[
                "scope_3_ghgp_c4_disclose_full"
            ]
            scope_3_ghgp_c4_exclusion_full = dataframes[
                "scope_3_ghgp_c4_exclusion_full"
            ]
            scope_3_ghgp_c4_change_full = dataframes[
                "scope_3_ghgp_c4_change_full"
            ]
            scope_3_ghgp_c5_full = dataframes["scope_3_ghgp_c5_full"]
            scope_3_ghgp_c5_method_full = dataframes[
                "scope_3_ghgp_c5_method_full"
            ]
            scope_3_ghgp_c5_disclose_full = dataframes[
                "scope_3_ghgp_c5_disclose_full"
            ]
            scope_3_ghgp_c5_exclusion_full = dataframes[
                "scope_3_ghgp_c5_exclusion_full"
            ]
            scope_3_ghgp_c5_change_full = dataframes[
                "scope_3_ghgp_c5_change_full"
            ]
            scope_3_ghgp_c6_full = dataframes["scope_3_ghgp_c6_full"]
            scope_3_ghgp_c6_method_full = dataframes[
                "scope_3_ghgp_c6_method_full"
            ]
            scope_3_ghgp_c6_disclose_full = dataframes[
                "scope_3_ghgp_c6_disclose_full"
            ]
            scope_3_ghgp_c6_exclusion_full = dataframes[
                "scope_3_ghgp_c6_exclusion_full"
            ]
            scope_3_ghgp_c6_change_full = dataframes[
                "scope_3_ghgp_c6_change_full"
            ]
            scope_3_ghgp_c7_full = dataframes["scope_3_ghgp_c7_full"]
            scope_3_ghgp_c7_method_full = dataframes[
                "scope_3_ghgp_c7_method_full"
            ]
            scope_3_ghgp_c7_disclose_full = dataframes[
                "scope_3_ghgp_c7_disclose_full"
            ]
            scope_3_ghgp_c7_exclusion_full = dataframes[
                "scope_3_ghgp_c7_exclusion_full"
            ]
            scope_3_ghgp_c7_change_full = dataframes[
                "scope_3_ghgp_c7_change_full"
            ]
            scope_3_ghgp_c8_full = dataframes["scope_3_ghgp_c8_full"]
            scope_3_ghgp_c8_method_full = dataframes[
                "scope_3_ghgp_c8_method_full"
            ]
            scope_3_ghgp_c8_disclose_full = dataframes[
                "scope_3_ghgp_c8_disclose_full"
            ]
            scope_3_ghgp_c8_exclusion_full = dataframes[
                "scope_3_ghgp_c8_exclusion_full"
            ]
            scope_3_ghgp_c8_change_full = dataframes[
                "scope_3_ghgp_c8_change_full"
            ]
            scope_3_ghgp_c9_full = dataframes["scope_3_ghgp_c9_full"]
            scope_3_ghgp_c9_method_full = dataframes[
                "scope_3_ghgp_c9_method_full"
            ]
            scope_3_ghgp_c9_disclose_full = dataframes[
                "scope_3_ghgp_c9_disclose_full"
            ]
            scope_3_ghgp_c9_exclusion_full = dataframes[
                "scope_3_ghgp_c9_exclusion_full"
            ]
            scope_3_ghgp_c9_change_full = dataframes[
                "scope_3_ghgp_c9_change_full"
            ]
            scope_3_ghgp_c10_full = dataframes["scope_3_ghgp_c10_full"]
            scope_3_ghgp_c10_method_full = dataframes[
                "scope_3_ghgp_c10_method_full"
            ]
            scope_3_ghgp_c10_disclose_full = dataframes[
                "scope_3_ghgp_c10_disclose_full"
            ]
            scope_3_ghgp_c10_exclusion_full = dataframes[
                "scope_3_ghgp_c10_exclusion_full"
            ]
            scope_3_ghgp_c10_change_full = dataframes[
                "scope_3_ghgp_c10_change_full"
            ]
            scope_3_ghgp_c11_full = dataframes["scope_3_ghgp_c11_full"]
            scope_3_ghgp_c11_method_full = dataframes[
                "scope_3_ghgp_c11_method_full"
            ]
            scope_3_ghgp_c11_disclose_full = dataframes[
                "scope_3_ghgp_c11_disclose_full"
            ]
            scope_3_ghgp_c11_exclusion_full = dataframes[
                "scope_3_ghgp_c11_exclusion_full"
            ]
            scope_3_ghgp_c11_change_full = dataframes[
                "scope_3_ghgp_c11_change_full"
            ]
            scope_3_ghgp_c12_full = dataframes["scope_3_ghgp_c12_full"]
            scope_3_ghgp_c12_method_full = dataframes[
                "scope_3_ghgp_c12_method_full"
            ]
            scope_3_ghgp_c12_exclusion_full = dataframes[
                "scope_3_ghgp_c12_exclusion_full"
            ]
            scope_3_ghgp_c12_disclose_full = dataframes[
                "scope_3_ghgp_c12_disclose_full"
            ]
            scope_3_ghgp_c12_change_full = dataframes[
                "scope_3_ghgp_c12_change_full"
            ]
            scope_3_ghgp_c13_full = dataframes["scope_3_ghgp_c13_full"]
            scope_3_ghgp_c13_method_full = dataframes[
                "scope_3_ghgp_c13_method_full"
            ]
            scope_3_ghgp_c13_disclose_full = dataframes[
                "scope_3_ghgp_c13_disclose_full"
            ]
            scope_3_ghgp_c13_exclusion_full = dataframes[
                "scope_3_ghgp_c13_exclusion_full"
            ]
            scope_3_ghgp_c13_change_full = dataframes[
                "scope_3_ghgp_c13_change_full"
            ]
            scope_3_ghgp_c14_full = dataframes["scope_3_ghgp_c14_full"]
            scope_3_ghgp_c14_method_full = dataframes[
                "scope_3_ghgp_c14_method_full"
            ]
            scope_3_ghgp_c14_disclose_full = dataframes[
                "scope_3_ghgp_c14_disclose_full"
            ]
            scope_3_ghgp_c14_exclusion_full = dataframes[
                "scope_3_ghgp_c14_exclusion_full"
            ]
            scope_3_ghgp_c14_change_full = dataframes[
                "scope_3_ghgp_c14_change_full"
            ]
            scope_3_ghgp_c15_full = dataframes["scope_3_ghgp_c15_full"]
            scope_3_ghgp_c15_method_full = dataframes[
                "scope_3_ghgp_c15_method_full"
            ]
            scope_3_ghgp_c15_disclose_full = dataframes[
                "scope_3_ghgp_c15_disclose_full"
            ]
            scope_3_ghgp_c15_exclusion_full = dataframes[
                "scope_3_ghgp_c15_exclusion_full"
            ]
            scope_3_ghgp_c15_change_full = dataframes[
                "scope_3_ghgp_c15_change_full"
            ]
            scope_3_ghgp_other_full = dataframes["scope_3_ghgp_other_full"]
            scope_3_ghgp_other_method_full = dataframes[
                "scope_3_ghgp_other_method_full"
            ]
            scope_3_ghgp_other_disclose_full = dataframes[
                "scope_3_ghgp_other_disclose_full"
            ]
            scope_3_ghgp_other_exclusion_full = dataframes[
                "scope_3_ghgp_other_exclusion_full"
            ]
            scope_3_ghgp_other_change_full = dataframes[
                "scope_3_ghgp_other_change_full"
            ]
            scope_3_iso_c3_full = dataframes["scope_3_iso_c3_full"]
            scope_3_iso_c3_method_full = dataframes[
                "scope_3_iso_c3_method_full"
            ]
            scope_3_iso_c3_disclose_full = dataframes[
                "scope_3_iso_c3_disclose_full"
            ]
            scope_3_iso_c3_exclusion_full = dataframes[
                "scope_3_iso_c3_exclusion_full"
            ]
            scope_3_iso_c3_change_full = dataframes[
                "scope_3_iso_c3_change_full"
            ]
            scope_3_iso_c4_full = dataframes["scope_3_iso_c4_full"]
            scope_3_iso_c4_method_full = dataframes[
                "scope_3_iso_c4_method_full"
            ]
            scope_3_iso_c4_disclose_full = dataframes[
                "scope_3_iso_c4_disclose_full"
            ]
            scope_3_iso_c4_exclusion_full = dataframes[
                "scope_3_iso_c4_exclusion_full"
            ]
            scope_3_iso_c4_change_full = dataframes[
                "scope_3_iso_c4_change_full"
            ]
            scope_3_iso_c5_full = dataframes["scope_3_iso_c5_full"]
            scope_3_iso_c5_method_full = dataframes[
                "scope_3_iso_c5_method_full"
            ]
            scope_3_iso_c5_disclose_full = dataframes[
                "scope_3_iso_c5_disclose_full"
            ]
            scope_3_iso_c5_exclusion_full = dataframes[
                "scope_3_iso_c5_exclusion_full"
            ]
            scope_3_iso_c5_change_full = dataframes[
                "scope_3_iso_c5_change_full"
            ]
            scope_3_iso_c6_full = dataframes["scope_3_iso_c6_full"]
            scope_3_iso_c6_method_full = dataframes[
                "scope_3_iso_c6_method_full"
            ]
            scope_3_iso_c6_disclose_full = dataframes[
                "scope_3_iso_c6_disclose_full"
            ]
            scope_3_iso_c6_exclusion_full = dataframes[
                "scope_3_iso_c6_exclusion_full"
            ]
            scope_3_iso_c6_change_full = dataframes[
                "scope_3_iso_c6_change_full"
            ]
            if self.company.company_type == "Financial":
                fe_aum_overview_full = fe_df_full["fe_aum_overview_full"]
                fe_aum_absolute_full = fe_df_full["fe_aum_absolute_full"]
                fe_aum_change_full = fe_df_full["fe_aum_change_full"]
                fe_aum_intensity_full = fe_df_full["fe_aum_intensity_full"]
                fe_aum_data_quality_full = fe_df_full[
                    "fe_aum_data_quality_full"
                ]
                fe_ge_overview_full = fe_df_full["fe_ge_overview_full"]
                fe_ge_absolute_full = fe_df_full["fe_ge_absolute_full"]
                fe_ge_change_full = fe_df_full["fe_ge_change_full"]
                fe_ge_intensity_full = fe_df_full["fe_ge_intensity_full"]
                fe_ge_data_quality_full = fe_df_full["fe_ge_data_quality_full"]
            metadata_fields_and_desc_full = pd.concat(
                [company_metadata_full],
                axis=0,
                ignore_index=True,
            )

            data_fields_emissions_tab_new = [
                scope_emissions_full,
                scope_exclusion_full,
                scope_emissions_change_full,
                scope_ghg_full,
                scope_ghg_other_full,
                scope_lb_emissions_full,
                scope_lb_exclusion_full,
                scope_lb_change_full,
                scope_mb_emissions_full,
                scope_mb_exclusion_full,
                scope_mb_change_full,
                scope_3_ghgp_emissions_full,
                scope_3_ghgp_c1_full,
                scope_3_ghgp_c1_method_full,
                scope_3_ghgp_c1_disclose_full,
                scope_3_ghgp_c1_exclusion_full,
                scope_3_ghgp_c1_change_full,
                scope_3_ghgp_c2_full,
                scope_3_ghgp_c2_method_full,
                scope_3_ghgp_c2_disclose_full,
                scope_3_ghgp_c2_exclusion_full,
                scope_3_ghgp_c2_change_full,
                scope_3_ghgp_c3_full,
                scope_3_ghgp_c3_method_full,
                scope_3_ghgp_c3_disclose_full,
                scope_3_ghgp_c3_exclusion_full,
                scope_3_ghgp_c3_change_full,
                scope_3_ghgp_c4_full,
                scope_3_ghgp_c4_method_full,
                scope_3_ghgp_c4_disclose_full,
                scope_3_ghgp_c4_exclusion_full,
                scope_3_ghgp_c4_change_full,
                scope_3_ghgp_c5_full,
                scope_3_ghgp_c5_method_full,
                scope_3_ghgp_c5_disclose_full,
                scope_3_ghgp_c5_exclusion_full,
                scope_3_ghgp_c5_change_full,
                scope_3_ghgp_c6_full,
                scope_3_ghgp_c6_method_full,
                scope_3_ghgp_c6_disclose_full,
                scope_3_ghgp_c6_exclusion_full,
                scope_3_ghgp_c6_change_full,
                scope_3_ghgp_c7_full,
                scope_3_ghgp_c7_method_full,
                scope_3_ghgp_c7_disclose_full,
                scope_3_ghgp_c7_exclusion_full,
                scope_3_ghgp_c7_change_full,
                scope_3_ghgp_c8_full,
                scope_3_ghgp_c8_method_full,
                scope_3_ghgp_c8_disclose_full,
                scope_3_ghgp_c8_exclusion_full,
                scope_3_ghgp_c8_change_full,
                scope_3_ghgp_c9_full,
                scope_3_ghgp_c9_method_full,
                scope_3_ghgp_c9_disclose_full,
                scope_3_ghgp_c9_exclusion_full,
                scope_3_ghgp_c9_change_full,
                scope_3_ghgp_c10_full,
                scope_3_ghgp_c10_method_full,
                scope_3_ghgp_c10_disclose_full,
                scope_3_ghgp_c10_exclusion_full,
                scope_3_ghgp_c10_change_full,
                scope_3_ghgp_c11_full,
                scope_3_ghgp_c11_method_full,
                scope_3_ghgp_c11_disclose_full,
                scope_3_ghgp_c11_exclusion_full,
                scope_3_ghgp_c11_change_full,
                scope_3_ghgp_c12_full,
                scope_3_ghgp_c12_method_full,
                scope_3_ghgp_c12_disclose_full,
                scope_3_ghgp_c12_exclusion_full,
                scope_3_ghgp_c12_change_full,
                scope_3_ghgp_c13_full,
                scope_3_ghgp_c13_method_full,
                scope_3_ghgp_c13_disclose_full,
                scope_3_ghgp_c13_exclusion_full,
                scope_3_ghgp_c13_change_full,
                scope_3_ghgp_c14_full,
                scope_3_ghgp_c14_method_full,
                scope_3_ghgp_c14_disclose_full,
                scope_3_ghgp_c14_exclusion_full,
                scope_3_ghgp_c14_change_full,
                scope_3_ghgp_c15_full,
                scope_3_ghgp_c15_method_full,
                scope_3_ghgp_c15_disclose_full,
                scope_3_ghgp_c15_exclusion_full,
                scope_3_ghgp_c15_change_full,
                scope_3_ghgp_other_full,
                scope_3_ghgp_other_method_full,
                scope_3_ghgp_other_disclose_full,
                scope_3_ghgp_other_exclusion_full,
                scope_3_ghgp_other_change_full,
                scope_3_iso_c3_full,
                scope_3_iso_c3_method_full,
                scope_3_iso_c3_disclose_full,
                scope_3_iso_c3_exclusion_full,
                scope_3_iso_c3_change_full,
                scope_3_iso_c4_full,
                scope_3_iso_c4_method_full,
                scope_3_iso_c4_disclose_full,
                scope_3_iso_c4_exclusion_full,
                scope_3_iso_c4_change_full,
                scope_3_iso_c5_full,
                scope_3_iso_c5_method_full,
                scope_3_iso_c5_disclose_full,
                scope_3_iso_c5_exclusion_full,
                scope_3_iso_c5_change_full,
                scope_3_iso_c6_full,
                scope_3_iso_c6_method_full,
                scope_3_iso_c6_disclose_full,
                scope_3_iso_c6_exclusion_full,
                scope_3_iso_c6_change_full,
            ]

            def first_notnull(series: pd.Series):
                non_null = series.dropna()
                if not non_null.empty:
                    return non_null.iloc[0]
                return np.nan

            original_order = pd.concat(
                data_fields_emissions_tab_new, axis=0, ignore_index=True
            )["Field Name"].drop_duplicates()
            order_dict = {name: i for i, name in enumerate(original_order)}

            if all(len(df) == 0 for df in data_fields_emissions_tab):
                data_fields_emissions_tab = list(data_fields_emissions_tab_new)

            else:
                for i, df in enumerate(data_fields_emissions_tab):
                    data_fields_emissions_tab[i] = (
                        data_fields_emissions_tab_new[i].merge(
                            df,
                            how="outer",
                        )
                    )
            date_fields_and_desc_full = (
                pd.concat(df for df in data_fields_emissions_tab)
                .groupby(["Field Name"], as_index=False)
                .agg(first_notnull)
                .assign(_position=lambda x: x["Field Name"].map(order_dict))
                .sort_values("_position")
                .drop("_position", axis=1)
                .reset_index(drop=True)
            )

            # iterate over scope_1_ghg Field Name column for
            # delete of suffix in scope_1_greenhouse_gas_other
            for index, row in date_fields_and_desc_full["Field Name"].items():
                # check if scope_1_greenhouse_gas_other is that key
                if row.startswith("scope_1_greenhouse_gas_other"):
                    # remove suffix from the row and update it
                    updated_row = row.rsplit("_", 1)[0]
                    date_fields_and_desc_full.at[index, "Field Name"] = (
                        updated_row
                    )
            if self.company.company_type == "Financial":
                fe_complete_full = pd.concat(
                    [
                        fe_aum_overview_full,
                        fe_aum_absolute_full,
                        fe_aum_change_full,
                        fe_aum_intensity_full,
                        fe_aum_data_quality_full,
                        fe_ge_overview_full,
                        fe_ge_absolute_full,
                        fe_ge_change_full,
                        fe_ge_intensity_full,
                        fe_ge_data_quality_full,
                    ]
                )

        with pd.ExcelWriter(excel_filename) as writer:  # pylint: disable=abstract-class-instantiated
            columns = ["Field Name", "Short Description"]
            columns_emissions_and_fe = [
                "Field Name",
                "Short Description",
                "Units",
            ]
            columns_targets = [
                "Target ID",
                "Target Name",
                "Field Name",
                "Short Description",
                "Units",
            ]
            rep_col_cm_year = 2
            rep_col_year = 3
            rep_col_tar_year = 5
            for year in range(2015, 2024):
                default = [
                    f"Value_{year}",
                    f"Source_{year}",
                    f"Last Updated_{year}",
                    f"Restated_{year}",
                ]
                # skip restated column for metadata dataframe
                default_metadata = [
                    col for col in default if "Restated" not in col
                ]
                columns.extend(default_metadata)
                columns_emissions_and_fe.extend(default)
                columns_targets.extend(default)
                # add company targets years to df
                pd.DataFrame({str(year): []}).to_excel(
                    writer,
                    sheet_name=CompaniesSheets.company_metadata.value,
                    startcol=rep_col_cm_year,
                    index=False,
                )
                # add company targets years to df
                pd.DataFrame({str(year): []}).to_excel(
                    writer,
                    sheet_name=CompaniesSheets.main_sheet.value,
                    startcol=rep_col_year,
                    index=False,
                )
                if self.company.company_type == "Financial":
                    # add company targets years to df
                    pd.DataFrame({str(year): []}).to_excel(
                        writer,
                        sheet_name=CompaniesSheets.financed_emissions_sheet.value,
                        startcol=rep_col_year,
                        index=False,
                    )
                rep_col_cm_year += 3
                rep_col_year += 4
                rep_col_tar_year += 4
            pd.DataFrame({col: [] for col in columns}).to_excel(
                writer,
                sheet_name=CompaniesSheets.company_metadata.value,
                startrow=1,
                index=False,
            )
            self._fill_no_disclosure_data(metadata_fields_and_desc_full)
            metadata_fields_and_desc_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.company_metadata.value,
                index=False,
                startrow=1,
            )
            pd.DataFrame(
                {col: [] for col in columns_emissions_and_fe}
            ).to_excel(
                writer,
                sheet_name=CompaniesSheets.main_sheet.value,
                startrow=1,
                index=False,
            )
            self._fill_no_disclosure_data(date_fields_and_desc_full)
            date_fields_and_desc_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.main_sheet.value,
                index=False,
                startrow=1,
            )
            if self.company.company_type == "Financial":
                pd.DataFrame(
                    {col: [] for col in columns_emissions_and_fe}
                ).to_excel(
                    writer,
                    sheet_name=CompaniesSheets.financed_emissions_sheet.value,
                    startrow=1,
                    index=False,
                )
                self._fill_no_disclosure_data(fe_complete_full)
                fe_complete_full.to_excel(
                    writer,
                    sheet_name=CompaniesSheets.financed_emissions_sheet.value,
                    index=False,
                    startrow=1,
                )
            # add assure verif to df
            assure_verif_full_list = []
            for year in range(2023, 2014, -1):
                assure_row = assure_verif_temp.get(year)
                if assure_row:
                    assure_verif_full_list += assure_row
                else:
                    default_assure = self._headers_assure_verif_config()
                    for key, value in default_assure.items():
                        assure_row = {
                            "Field Name": key,
                            "Short Description": value,
                            "Reporting Year": year,
                            "Value": NullTypeState.LONG_DASH.value,
                            "Source": None,
                            "Last Updated": "blank",
                            "Restated": EN_DASH,
                        }
                        assure_verif_full_list.append(assure_row)
            # reverse assure verif full list
            assure_verif_full = pd.DataFrame(assure_verif_full_list)
            self._fill_no_disclosure_data(assure_verif_full)
            assure_verif_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.assure_verif.value,
                index=False,
            )
            # add restatements to df
            if restatements_df.empty:
                restatement_default = {
                    "Reporting Year": "No Restatement Data Available.",
                    "Data Model": None,
                    "Restated Field Name": None,
                    "Restated Short Description": None,
                    "Reporting Date": None,
                    "Field Value": None,
                    "Field Units": None,
                    "Restatement Rationale": None,
                    "Source": None,
                }
                restatements_df = pd.DataFrame([restatement_default])
                restatements_df.to_excel(
                    writer,
                    sheet_name=CompaniesSheets.restatement_sheet.value,
                    index=False,
                )
            else:
                restatements_df.to_excel(
                    writer,
                    sheet_name=CompaniesSheets.restatement_sheet.value,
                    index=False,
                )
            # add targets to dataframe
            if len(targets_full_list) == 0:
                targets_full_list = self._no_data_default_headers(
                    worksheet=CompaniesSheets.emissions_reduction_targets,
                    plural=True,
                )
            targets_full = pd.DataFrame(targets_full_list)
            self._fill_no_disclosure_data(targets_full)
            targets_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.emissions_reduction_targets.value,
                index=False,
            )
            if len(targets_progress_full_list) == 0:
                targets_progress_full_list = self._no_data_default_headers(
                    worksheet=CompaniesSheets.targets_progress_sheet,
                    plural=True,
                )
            targets_progress_full = pd.DataFrame(targets_progress_full_list)
            self._fill_no_disclosure_data(targets_progress_full)
            targets_progress_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.targets_progress_sheet.value,
                index=False,
            )
            if len(validation_full_list) == 0:
                validation_full_list = self._no_data_default_headers(
                    worksheet=CompaniesSheets.validation_sheet
                )
            # add validation to df
            validation_full = pd.DataFrame(validation_full_list)
            self._fill_no_disclosure_data(validation_full)
            validation_full.to_excel(
                writer,
                sheet_name=CompaniesSheets.validation_sheet.value,
                index=False,
            )
        year_sheets = [
            CompaniesSheets.company_metadata.value,
            CompaniesSheets.main_sheet.value,
            CompaniesSheets.financed_emissions_sheet.value,
        ]

        wb = load_workbook(excel_filename)
        ws_order = [
            CompaniesSheets.company_metadata.value,
            CompaniesSheets.main_sheet.value,
            CompaniesSheets.financed_emissions_sheet.value,
            CompaniesSheets.assure_verif.value,
            CompaniesSheets.restatement_sheet.value,
            CompaniesSheets.emissions_reduction_targets.value,
            CompaniesSheets.targets_progress_sheet.value,
            CompaniesSheets.validation_sheet.value,
        ]
        if self.company.company_type != "Financial":
            year_sheets.remove(CompaniesSheets.financed_emissions_sheet.value)
            ws_order.remove(CompaniesSheets.financed_emissions_sheet.value)
        wb._sheets = [
            wb[sheet] for sheet in ws_order if sheet in wb.sheetnames
        ]
        left_alignment = Alignment(horizontal="left")
        normal_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None),
        )
        for ws in wb.worksheets:
            if ws.title in year_sheets:
                # strip years from headers
                for cell in ws[2]:
                    if isinstance(cell.value, str) and "_" in cell.value:
                        cell.value = cell.value.split("_")[0]
                    cell.alignment = left_alignment
                first_populated_col = None
                for cell in ws[1]:
                    if cell.value is not None:
                        first_populated_col = cell.column
                        break
                # fill blanks with dash
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=ws.max_row,
                    min_col=first_populated_col,
                    max_col=ws.max_column,
                ):
                    for cell in row:
                        if cell.value is None:
                            cell.value = EN_DASH
                            pass
                        # remove blank with None value for last_updated cols
                        elif cell.value == "blank":
                            cell.value = None

            if ws.title in [
                CompaniesSheets.assure_verif.value,
                CompaniesSheets.emissions_reduction_targets.value,
                CompaniesSheets.targets_progress_sheet.value,
                CompaniesSheets.validation_sheet.value,
                CompaniesSheets.restatement_sheet.value,
            ]:
                if ws.max_row > 2:
                    for row in ws.iter_rows(
                        min_row=2,
                        max_row=ws.max_row,
                        min_col=6,
                        max_col=ws.max_column,
                    ):
                        for cell in row:
                            if cell.value is None:
                                cell.value = EN_DASH
                            # remove blank with None value for last_updated cols
                            elif cell.value == "blank":
                                cell.value = None
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = normal_border
                    if cell.row == 1:
                        cell.alignment = left_alignment
        wb.save(excel_filename)
        return excel_filename
